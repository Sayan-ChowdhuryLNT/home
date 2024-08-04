import hashlib
import random
import sqlite3
import string
import time
import openpyxl
import win32com
from Mainapp.auth import db, Project
from Mainapp.auth import get_email
import datetime
from flask import Blueprint, jsonify, render_template, send_file,session,request
from datetime import datetime, timedelta
import time
import math
pii = math.pi
import shutil
import time
import os, os.path
from math import *
import pandas as pd
from Mainapp.documentation import *
import shutil
import os
import schedule
import win32com.client
from werkzeug.security import generate_password_hash, check_password_hash

import pythoncom
db_bp= Blueprint("database", __name__, template_folder="templates")
master_emails= ["sayan.chowdhury@lntecc.com","balajirajadurai@lntecc.com","niraj.nair@lntecc.com","karthiga@lntecc.com","sudripta.misra@lntecc.com",'srb@lntecc.com',"rpvenkatesh@lntecc.com","bnk@lntecc.com"]

# The above code snippet is defining a Blueprint named "database"
# with a specified template folder "templates". Additionally, there is a list called "master_emails"
# which contains email addresses of individuals.




def rev_d(original_date):
    REV_D_DOCSTRING
    try:
        date_object = datetime.strptime(original_date, "%Y-%m-%d")
        reversed_date = date_object.strftime("%d-%m-%Y")

        return reversed_date
    except ValueError as e:

        return f"Error: {str(e)}"
    

def fetch_dropdown_options(dropdown_name):
    FETCH_DROPDOWN_OPTIONS_DOCSTRING
    conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')
    cursor = conn.cursor()

    cursor.execute(f'SELECT DISTINCT {dropdown_name} FROM project ORDER BY {dropdown_name} ASC;')
    options = [row[0] for row in cursor.fetchall()]


    conn.close()

    return options



    
    
def cycle_time(dates):
    CYCLE_TIME_DOCSTRING
    total_days_difference = 0
    filled_rows_count = 0
    for date in dates:
        date_str1, date_str2 = date
        if date_str1 and date_str2:
            if date_str1.strip() and date_str2.strip():  
                date1 = datetime.strptime(date_str1, '%d.%m.%y')
                date2 = datetime.strptime(date_str2, '%d.%m.%y')
                difference = date2 - date1
                days_difference = difference.days
                total_days_difference += days_difference
                filled_rows_count += 1

    if filled_rows_count > 0:
        average_days_difference = round(total_days_difference / filled_rows_count) 
        return average_days_difference


def fetch_discipline_options(dropdown_name):
    FETCH_DISCIPLINE_OPTIONS_DOCSTRING
    conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')
    cursor = conn.cursor()

    cursor.execute(f'SELECT DISTINCT {dropdown_name} FROM pr_data ORDER BY {dropdown_name} ASC;')
    doptions = [row[0] for row in cursor.fetchall()]
    print(doptions)
    doptions=[item for item in doptions if item != '']
    print(doptions)
    conn.close()
    return doptions      





def elapsed_time(dates,meta_data_status):
        ELAPSED_TIME_DOCSTRING
        conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')
        cursor = conn.cursor()
        # print(meta_data_status)
        status=[]
        pr_code=[]
        for metadata in meta_data_status:
            status_meta, pr_code_meta=metadata
            status.append(status_meta)
            pr_code.append(pr_code_meta)
            
        session['item_vendorCode_query']=f'''SELECT item_vendor_code FROM pr_data'''
        cursor.execute(session['item_vendorCode_query'])
        session['item_vendorCode_query_result'] = cursor.fetchall()
        df_item_vendor_code=pd.DataFrame(session['item_vendorCode_query_result'])   
        print(df_item_vendor_code.shape[0])
        item_code_result_list=[]
        
        pr_code_list=[]
        session['statuses_to_check_elapsed'] = [
        "Awaiting post-offer from vendor",
        "Post-order pending with EDRC",
        "Post-order comments issued",
        "Post-order technically cleared",
        "Approved by Client"
    ]
        for j in range (len(status)):
            if status[j] in session['statuses_to_check_elapsed']:
                pr_code_list.append(pr_code[j])
            
        days_list=[]
        i=0
       
        for date_row in dates:
            # if status[i] in session['statuses_to_check_elapsed']:
            #     pr_code_list.append(pr_code[i])
                
            if status[i]!= "Approved by Client" or status[i]!= "Enquiry yet to float" or status[i]!= "Regret offer" or status[i]!= "Sent to BU Operations" or status[i]!= "Pre-order technically rejected":
                cleaned_dates = [date for date in date_row if date is not None and date.strip() != "" and len(date)<=8 and date!='19.0.24']
                if cleaned_dates:
                    datetime_dates = [datetime.strptime(date, '%d.%m.%y') for date in cleaned_dates]
                    latest_date = max(datetime_dates)
                    difference = (datetime.now() - latest_date).days
                else:
                    difference="NA"
            
            if status[i]== "Approved by Client" or status[i]== "Regret offer" or status[i]== "Pre-order technically rejected":
                difference="PR Closed"
            if status[i]== "Enquiry yet to float":
                difference="NA"
            if status[i] == "Sent to BU Operations":
                if pr_code[i] in pr_code_list:
                    
                
                # # print(itemVendorCodeTocheck)
                # if itemVendorCodeTocheck in item_code_result_list:
                #     # print('yes')
                    
                    difference="PR Closed"
                    
                else:
                    # print('No')
                    cleaned_dates = [date for date in date_row if date is not None and date.strip() != "" and len(date)<=8]
                    if cleaned_dates:
                        datetime_dates = [datetime.strptime(date, '%d.%m.%y') for date in cleaned_dates]
                        latest_date = max(datetime_dates)
                        difference = (datetime.now() - latest_date).days
                    else:
                        difference="NA"
                
            days_list.append(difference)
            i+=1
 
            

        return days_list




def calculate_average_days(date):
    CALCULATE_AVERAGE_DAYS_DOCSTRING
    session['total_days_difference_edrc_pre_A'] = 0
    session['total_days_difference_edrc_pre_B'] = 0
    session['total_days_difference_edrc_pre_C'] = 0
    session['total_days_difference_edrc_pre_D'] = 0
    session['total_days_difference_edrc_pre_E'] = 0
    session['total_days_difference_edrc_pre_F'] = 0
    
    session['total_days_difference_vendor_pre_A'] = 0
    session['total_days_difference_vendor_pre_B'] = 0
    session['total_days_difference_vendor_pre_C'] = 0
    session['total_days_difference_vendor_pre_D'] = 0
    session['total_days_difference_vendor_pre_F'] = 0
    session['total_days_difference_vendor_pre_E'] = 0
    
    
    session['total_days_difference_edrc_post_A'] = 0
    session['total_days_difference_edrc_post_B'] = 0
    
    session['total_days_difference_vendor_post_A'] = 0
    session['total_days_difference_vendor_post_B'] = 0
    

    session['filled_pairs_count_edrc_pre'] = 0
    session['filled_pairs_count_vendor_pre'] = 0
    session['filled_pairs_count_edrc_post'] = 0
    session['filled_pairs_count_vendor_post'] = 0
    
    for dates in date:
        session['counta_vendor_post']=0
        session['counta_vendor_pre']=0
        
        session['counta_edrc_post']=0
        session['counta_edrc_pre']=0
        
        session['EnquirySent'], session['OfferRcvA'], session['CommentsSentA'], session['OfferRcvB'], session['CommentsSentB'], session['OfferRcvC'], session['CommentsSentC'], session['OfferRcvD'], session['CommentsSentD'], session['OfferRcvE'], session['CommentsSentE'], session['DocfromVendorPRE'], session['SentSCM'], session['POissued'], session['PostRCV'], session['CommentsEDRCPOST'], session['DocfromVendorPOST'], session['MFCissued'] = dates

        # session['prissuedDate'] = datetime.strptime(session['prissuedDate'], '%d.%m.%y') if session['prissuedDate'] is not None and len(session['prissuedDate']) == 8 else None
        session['EnquirySent'] = datetime.strptime(session['EnquirySent'], '%d.%m.%y') if session['EnquirySent'] is not None and len(session['EnquirySent']) == 8 else None
        session['OfferRcvA'] = datetime.strptime(session['OfferRcvA'], '%d.%m.%y') if session['OfferRcvA'] is not None and len(session['OfferRcvA']) == 8 else None
        session['CommentsSentA'] = datetime.strptime(session['CommentsSentA'], '%d.%m.%y') if session['CommentsSentA'] is not None and len(session['CommentsSentA']) == 8 else None
        session['OfferRcvB'] = datetime.strptime(session['OfferRcvB'], '%d.%m.%y') if session['OfferRcvB'] is not None and len(session['OfferRcvB']) == 8 else None
        session['CommentsSentB'] = datetime.strptime(session['CommentsSentB'], '%d.%m.%y') if session['CommentsSentB'] is not None and len(session['CommentsSentB']) == 8 else None
        session['OfferRcvC'] = datetime.strptime(session['OfferRcvC'], '%d.%m.%y') if session['OfferRcvC'] is not None and len(session['OfferRcvC']) == 8 else None
        session['CommentsSentC'] = datetime.strptime(session['CommentsSentC'], '%d.%m.%y') if session['CommentsSentC'] is not None and len(session['CommentsSentC']) == 8 else None
        session['OfferRcvD'] = datetime.strptime(session['OfferRcvD'], '%d.%m.%y') if session['OfferRcvD'] is not None and len(session['OfferRcvD']) == 8 else None
        session['CommentsSentD'] = datetime.strptime(session['CommentsSentD'], '%d.%m.%y') if session['CommentsSentD'] is not None and len(session['CommentsSentD']) == 8 else None
        session['OfferRcvE'] = datetime.strptime(session['OfferRcvE'], '%d.%m.%y') if session['OfferRcvE'] is not None and len(session['OfferRcvE']) == 8 else None
        session['CommentsSentE'] = datetime.strptime(session['CommentsSentE'], '%d.%m.%y') if session['CommentsSentE'] is not None and len(session['CommentsSentE']) == 8 else None
        session['DocfromVendorPRE'] = datetime.strptime(session['DocfromVendorPRE'], '%d.%m.%y') if session['DocfromVendorPRE'] is not None and len(session['DocfromVendorPRE']) == 8 else None
        session['SentSCM'] = datetime.strptime(session['SentSCM'], '%d.%m.%y') if session['SentSCM'] is not None and len(session['SentSCM']) == 8 else None
        session['POissued'] = datetime.strptime(session['POissued'], '%d.%m.%y') if session['POissued'] is not None and len(session['POissued']) == 8 else None
        session['PostRCV'] = datetime.strptime(session['PostRCV'], '%d.%m.%y') if session['PostRCV'] is not None and len(session['PostRCV']) == 8 else None
        session['CommentsEDRCPOST'] = datetime.strptime(session['CommentsEDRCPOST'], '%d.%m.%y') if session['CommentsEDRCPOST'] is not None and len(session['CommentsEDRCPOST']) == 8 else None
        session['DocfromVendorPOST'] = datetime.strptime(session['DocfromVendorPOST'], '%d.%m.%y') if session['DocfromVendorPOST'] is not None and len(session['DocfromVendorPOST']) == 8 else None
        session['MFCissued'] = datetime.strptime(session['MFCissued'], '%d.%m.%y') if session['MFCissued'] is not None and len(session['MFCissued']) == 8 else None

        
        # Calculate days differences for EDRC 
        if session['OfferRcvA'] and session['CommentsSentA']:
            session['WITH_EDRC_A'] = (session['CommentsSentA'] - session['OfferRcvA']).days
            session['total_days_difference_edrc_pre_A'] += session['WITH_EDRC_A']
            session['counta_edrc_pre']+=1
            

        if session['OfferRcvB'] and session['CommentsSentB']:
            session['WITH_EDRC_B'] = (session['CommentsSentB'] - session['OfferRcvB']).days
            session['total_days_difference_edrc_pre_B'] += session['WITH_EDRC_B']
            session['counta_edrc_pre']+=1
            
            

        if session['OfferRcvC'] and session['CommentsSentC']:
            session['WITH_EDRC_C'] = (session['CommentsSentC'] - session['OfferRcvC']).days
            session['total_days_difference_edrc_pre_C'] += session['WITH_EDRC_C']
            session['counta_edrc_pre']+=1
            
           

        if session['OfferRcvD'] and session['CommentsSentD']:
            session['WITH_EDRC_D'] = (session['CommentsSentD'] - session['OfferRcvD']).days
            session['total_days_difference_edrc_pre_D'] += session['WITH_EDRC_D']
            session['counta_edrc_pre']+=1
            
        

        if session['OfferRcvE'] and session['CommentsSentE']:
            session['WITH_EDRC_E'] = (session['CommentsSentE'] - session['OfferRcvE']).days
            session['total_days_difference_edrc_pre_E'] += session['WITH_EDRC_E']
            session['counta_edrc_pre']+=1
            
         

        if session['SentSCM'] and session['DocfromVendorPRE']:
            session['WITH_EDRC_F'] = (session['SentSCM'] - session['DocfromVendorPRE']).days
            session['total_days_difference_edrc_pre_F'] += session['WITH_EDRC_F']
            session['counta_edrc_pre']+=1


        # Calculate days differences for Vend 

        if session['EnquirySent'] and session['OfferRcvA']:
            session['WITH_VENDOR_A'] = (session['OfferRcvA'] - session['EnquirySent']).days
            session['total_days_difference_vendor_pre_A'] += session['WITH_VENDOR_A']
            session['counta_vendor_pre']+=1
            
            

        if session['CommentsSentA'] and session['OfferRcvB']:
            session['WITH_VENDOR_B'] = (session['OfferRcvB'] - session['CommentsSentA']).days
            session['total_days_difference_vendor_pre_B'] += session['WITH_VENDOR_B']
            session['counta_vendor_pre']+=1
            
            

        if session['CommentsSentB'] and session['OfferRcvC']:
            session['WITH_VENDOR_C'] = (session['OfferRcvC'] - session['CommentsSentB']).days
            session['total_days_difference_vendor_pre_C'] += session['WITH_VENDOR_C']
            session['counta_vendor_pre']+=1
            
      

        if session['CommentsSentC'] and session['OfferRcvD']:
            session['WITH_VENDOR_D'] = (session['OfferRcvD'] - session['CommentsSentC']).days
            session['total_days_difference_vendor_pre_D'] += session['WITH_VENDOR_D']
            session['counta_vendor_pre']+=1
            
           

        if session['CommentsSentD'] and session['OfferRcvE']:
            session['WITH_VENDOR_E'] = (session['OfferRcvE'] - session['CommentsSentD']).days
            session['total_days_difference_vendor_pre_E'] += session['WITH_VENDOR_E']
            session['counta_vendor_pre']+=1
            
            

        if session['DocfromVendorPRE'] and session['CommentsSentE']:
            session['WITH_VENDOR_F'] = (session['DocfromVendorPRE'] - session['CommentsSentE']).days
            session['total_days_difference_vendor_pre_F'] += session['WITH_VENDOR_F']
            session['counta_vendor_pre']+=1
            
          

        # Calculate days differences for EDRC post-process
        if session['CommentsEDRCPOST'] and session['PostRCV']:
            session['WITH_EDRC_POST_A'] = (session['CommentsEDRCPOST'] - session['PostRCV']).days
            session['total_days_difference_edrc_post_A'] += session['WITH_EDRC_POST_A']
            session['counta_edrc_post']+=1
            
        

        if session['MFCissued'] and session['DocfromVendorPOST']:
            session['WITH_EDRC_POST_B'] = (session['MFCissued'] - session['DocfromVendorPOST']).days
            session['total_days_difference_edrc_post_B'] += session['WITH_EDRC_POST_B']
            session['counta_edrc_post']+=1
            
            

        # Calculate days differences for Vendor post-process
        if session['POissued'] and session['PostRCV']:
            session['WITH_VENDOR_POST_A'] = (session['PostRCV'] - session['POissued']).days
            session['total_days_difference_vendor_post_A'] += session['WITH_VENDOR_POST_A']
            session['counta_vendor_post']+=1
            
            

        if session['DocfromVendorPOST'] and session['CommentsEDRCPOST']:
            session['WITH_VENDOR_POST_B'] = (session['DocfromVendorPOST'] - session['CommentsEDRCPOST']).days
            session['total_days_difference_vendor_post_B'] += session['WITH_VENDOR_POST_B']
            session['counta_vendor_post']+=1
   
        
        if session['counta_edrc_pre']>0:  
            session['filled_pairs_count_edrc_pre'] += 1
        if session['counta_edrc_post']>0:  
            session['filled_pairs_count_edrc_post'] += 1  
        if session['counta_vendor_pre']>0:     
            session['filled_pairs_count_vendor_pre'] += 1
        if session['counta_vendor_post']>0:  
            session['filled_pairs_count_vendor_post'] += 1
            
            
    # Calculate averages
    session['average_days_EDRC_PRE'] = round((session['total_days_difference_edrc_pre_A']+session['total_days_difference_edrc_pre_B']+session['total_days_difference_edrc_pre_C']+session['total_days_difference_edrc_pre_D']+session['total_days_difference_edrc_pre_E']+session['total_days_difference_edrc_pre_F']) / session['filled_pairs_count_edrc_pre']) if session['filled_pairs_count_edrc_pre'] > 0 else 0
    session['average_days_VENDOR_PRE'] = round((session['total_days_difference_vendor_pre_A']+session['total_days_difference_vendor_pre_B']+session['total_days_difference_vendor_pre_C']+session['total_days_difference_vendor_pre_D']+session['total_days_difference_vendor_pre_E']+session['total_days_difference_vendor_pre_F']) / session['filled_pairs_count_vendor_pre']) if session['filled_pairs_count_vendor_pre'] > 0 else 0
    session['average_days_EDRC_POST'] = round((session['total_days_difference_edrc_post_A']+session['total_days_difference_edrc_post_B']) / session['filled_pairs_count_edrc_post']) if session['filled_pairs_count_edrc_post'] > 0 else 0
    session['average_days_VENDOR_POST'] = round((session['total_days_difference_vendor_post_A']+session['total_days_difference_vendor_post_B'])/ session['filled_pairs_count_vendor_post']) if session['filled_pairs_count_vendor_post'] > 0 else 0

    return session['average_days_EDRC_PRE'], session['average_days_VENDOR_PRE'], session['average_days_EDRC_POST'], session['average_days_VENDOR_POST']



@db_bp.route('/download_template')
def download_template():
    DOWNLOAD_TEMPLATE_DOCSTRING
    template_file_path = r'C:\Users\20323801\Desktop\VendorDesk Dev\WWWSBG\Mainapp\Detailed_PR_status.xlsx'

    return send_file(template_file_path, as_attachment=True)



@db_bp.route('/exportxl', methods=['POST'])
def exportxl():
    EXPORTXL_DOCSTRING
    session["xldata"] = request.get_json() 
    df = pd.DataFrame(session["xldata"])
    APP_ROOT = os.path.dirname(os.path.abspath(__file__))
    xl_file_path = os.path.join(APP_ROOT, 'Detailed_PR_Status.xlsx')
    xl_file_path1 = os.path.join(APP_ROOT, 'Detailed_PR_Status-1.xlsx')
    wb = openpyxl.load_workbook(xl_file_path)
    ws = wb.active
    folpath = "C:\\WetdeskDownloads\\export\\status"
    if "user" in session:
        userr = session["user"]
        session["folpath_user"] = folpath + "\\" + str(userr)
        

    if not os.path.exists(session["folpath_user"]):
            os.makedirs(session["folpath_user"])
            print("NEW USER PATH: ",session["folpath_user"])
    file_list = os.listdir(session["folpath_user"])
    print(xl_file_path)
    
    for r_idx, row in enumerate(df.values.tolist(), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(xl_file_path1)
    for file_name in file_list:
        file_path = os.path.join(session["folpath_user"], file_name)
        try:
            # Delete the xl_file
            os.remove(file_path)
            print(f"Deleted: {file_path}")
        except Exception as e:
            print(f"Error deleting {file_path}: {e}")

    source_path = xl_file_path1
    print(source_path)
    session["excel_destination_path"] = session["folpath_user"] + "\\" + "Detailed_PR_Status.xlsx"
    print(session["excel_destination_path"])

    if os.path.isfile(source_path):
        shutil.copy(source_path, session["excel_destination_path"])
    else:
        print(f'Error: {source_path} does not exist')

    t = os.path.getctime(session["excel_destination_path"])
    t_str = time.ctime(t)    
    t_obj = time.strptime(t_str)  
    form_t = time.strftime("Detailed_PR_Status-%Y-%m-%d%H:%M:%S", t_obj)    
    form_t = form_t.replace(":", "")   
    form_t = form_t.replace("-", "") 
    os.rename(session["excel_destination_path"], os.path.split(session["excel_destination_path"])[0] + '/' + form_t + os.path.splitext(session["excel_destination_path"])[1])
    session["excel_destination_path"] = session["excel_destination_path"].replace("Detailed_PR_Status",form_t)

    print(session["excel_destination_path"])
    
    return send_file(session["excel_destination_path"], as_attachment=True)




def chart(discipline,project_name, bu, sbg, ic):
    CHART_DOCSTRING
    conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')
    cursor = conn.cursor()

    session["base_query1"] = '''
        WITH MaxStage AS (
            SELECT
                pr_data.item_code AS max_item_code,  
                MAX(CASE
                        WHEN status IN ('Approved by Client') THEN 4
                        WHEN status IN ('Post-order comments issued', 'Awaiting post-offer from vendor', 'Post-order pending with EDRC', 'Post-order technically cleared') THEN 3
                        WHEN status IN ('Sent to BU Operations') THEN 2
                        WHEN status IN  ('Awaiting pre-offer from vendor', 'Pre-order pending with EDRC', 'Pre-order comments issued',
                        'Pre-order technically cleared', 'Pre-order technically rejected') THEN 1
                        WHEN status IN ('Enquiry yet to float') THEN 0
                        
                    END) AS max_priority
            FROM pr_data
            GROUP BY pr_data.item_code
        )
    '''

    session["base_query1A"] = '''
        SELECT
            discipline,
            COUNT(DISTINCT CASE WHEN max_priority = 4 THEN max_item_code END) AS CompletedCount,
            COUNT(DISTINCT CASE WHEN max_priority = 3 THEN max_item_code END) AS PostOrderStageCount,
            COUNT(DISTINCT CASE WHEN max_priority = 2 THEN max_item_code END) AS PreOrderCompletedCount,
            COUNT(DISTINCT CASE WHEN max_priority = 1 THEN max_item_code END) AS PreOrderStageCount,
            COUNT(DISTINCT CASE WHEN max_priority = 0 THEN max_item_code END) AS NotYetStartedCount
        FROM pr_data
        JOIN MaxStage ON pr_data.item_code = MaxStage.max_item_code  
    '''
 
    session["base_query2"]='''SELECT  project_name,discipline,equipment,item_code,item_vendor_code,vendor_name,
                            status,elapsed_days,remarks,vendor_ph,vendor_mail,pr_issue_date_pre,pr_issue_date_final,offer_recv_date,
                            comments_sent_pre,offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                            offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,
                            tc_meeting_vendor_pre,final_vendor_doc_pre,sent_to_scm,priority,po_issued_date,
                            finalized_vendor,post_doc_rcv_dt,comments_sent_post,tc_meeting_vendor_post,
                            final_vendor_doc_post,mfc_issued_date,target_completion_date,  target_completion_date_po,target_completion_date_post,last_modified_by
                            FROM pr_data
    '''

    
    session["base_query3"] = '''
        SELECT
            COUNT(DISTINCT CASE WHEN status IN (
                'Awaiting pre-offer from vendor',
                'Pre-order comments issued'
            ) THEN item_vendor_code END) AS PendingWithVendorPRE,
            COUNT(DISTINCT CASE WHEN status IN ( 
                'Pre-order pending with EDRC'
            ) THEN item_vendor_code END) AS PendingWithEDRCPRE,
            COUNT(DISTINCT CASE WHEN status IN (
                'Pre-order technically cleared',
                'Pre-order technically rejected',
                'Regret offer',
                'Sent to BU Operations',
                'Approved by Client',
                'Awaiting post-offer from vendor',
                'Post-order comments issued',
                'Post-order pending with EDRC',
                'Post-order technically cleared'
                
            ) THEN item_vendor_code END) AS Cleared
              FROM pr_data
    '''
       
    session["base_query4"] = '''
        SELECT
             COUNT(DISTINCT CASE WHEN status IN (
                 'Awaiting post-offer from vendor',
                 'Post-order comments issued'
             ) THEN item_vendor_code END) AS PendingWithVendorPOST,
              COUNT(DISTINCT CASE WHEN status IN (
                 'Post-order pending with EDRC'
             ) THEN item_vendor_code END) AS PendingWithEDRCPOST,
              COUNT(DISTINCT CASE WHEN status IN (
                 'Post-order technically cleared',
                 'Approved by Client'
             ) THEN item_vendor_code END) AS Completed
              FROM pr_data
    '''
            
    session["base_query5"] = '''
        SELECT
               COUNT(DISTINCT CASE WHEN status IN (
                'Pre-order technically cleared',
                'Pre-order technically rejected',
                'Regret offer',
                'Sent to BU Operations',
                'Approved by Client',
                'Awaiting post-offer from vendor',
                'Post-order comments issued',
                'Post-order pending with EDRC',
                'Post-order technically cleared'
                
            ) THEN item_code END) AS  total_pr_cleared,
              COUNT(DISTINCT CASE WHEN status IN (
                 'Post-order technically cleared',
                 'Approved by Client'
                 
             ) THEN item_code END) AS ClearedPO
              FROM pr_data
    '''
    session['float_days_query']='''SELECT pr_issue_date_final,sent_to_scm FROM pr_data'''
    session['pre_dates_query']='''SELECT sent_to_scm,po_issued_date FROM pr_data'''
    session['post_dates_query']='''SELECT po_issued_date,mfc_issued_date FROM pr_data'''


    session['average_time']='''SELECT pr_issue_date_final,offer_recv_date,
                            comments_sent_pre,offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                            offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,
                            final_vendor_doc_pre,sent_to_scm,po_issued_date,
                            post_doc_rcv_dt,comments_sent_post,
                            final_vendor_doc_post,mfc_issued_date FROM pr_data '''
    session['status_']='''SELECT status,item_code FROM pr_data'''
    session['process_ppi_query_1']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Awaiting pre-offer from vendor' AND discipline='Process'  '''
    session['process_ppi_query_2']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Pre-order technically cleared' AND discipline='Process'  '''
    session['process_ppi_query_3']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Awaiting post-offer from vendor' AND discipline='Process'  '''
    session['process_ppi_query_4']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Post-order technically cleared' AND discipline='Process'  '''
    session['process_ppi_query_5']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Approved by Client' AND discipline='Process'  '''
    
    session['Mechanical_ppi_query_1']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Awaiting pre-offer from vendor' AND discipline='Mechanical'  '''
    session['Mechanical_ppi_query_2']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Pre-order technically cleared' AND discipline='Mechanical'  '''
    session['Mechanical_ppi_query_3']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Awaiting post-offer from vendor' AND discipline='Mechanical'  '''
    session['Mechanical_ppi_query_4']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Post-order technically cleared' AND discipline='Mechanical'  '''
    session['Mechanical_ppi_query_5']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Approved by Client' AND discipline='Mechanical'  '''
    
    session['Electrical_ppi_query_1']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Awaiting pre-offer from vendor' AND discipline='Electrical'  '''
    session['Electrical_ppi_query_2']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Pre-order technically cleared' AND discipline='Electrical'  '''
    session['Electrical_ppi_query_3']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Awaiting post-offer from vendor' AND discipline='Electrical'  '''
    session['Electrical_ppi_query_4']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Post-order technically cleared' AND discipline='Electrical'  '''
    session['Electrical_ppi_query_5']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Approved by Client' AND discipline='Electrical'  '''
    
    session['Instrumentation_ppi_query_1']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Awaiting pre-offer from vendor' AND discipline='Instrumentation'  '''
    session['Instrumentation_ppi_query_2']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Pre-order technically cleared' AND discipline='Instrumentation'  '''
    session['Instrumentation_ppi_query_3']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Awaiting post-offer from vendor' AND discipline='Instrumentation'  '''
    session['Instrumentation_ppi_query_4']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Post-order technically cleared' AND discipline='Instrumentation'  '''
    session['Instrumentation_ppi_query_5']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data WHERE STATUS = 'Approved by Client' AND discipline='Instrumentation'  '''
    session['total_pr_no_query']='''SELECT COUNT(DISTINCT item_vendor_code) FROM pr_data '''
    
    
    
    session["conditions"] = []
    session["parameters"] = []

    if project_name != 'Project' and project_name != 'ALL' and project_name is not None:
        print(project_name)
        session['project'] = project_name
        
        cursor.execute('SELECT job_code FROM pr_data WHERE project_name=?',(session['project'],))
        
        job_code =cursor.fetchone()[0]
            
        session["conditions"].append("job_code = ?")
        session["parameters"].append(job_code)

    if ic != 'IC' and ic != 'ALL'   and ic is not None:
        session["conditions"].append("ic = ?")
        session["parameters"].append(ic)

    if sbg != 'sbg' and  sbg != 'ALL'  and sbg is not None:
        session["conditions"].append("sbg = ?")
        session["parameters"].append(sbg)

    if bu != 'BU' and  bu != 'ALL'  and bu is not None:
        session["conditions"].append("bu = ?")
        session["parameters"].append(bu)
    if discipline!= 'Discipline' and  discipline != 'ALL'  and discipline is not None:
        session["conditions"].append("discipline = ?")
        session["parameters"].append(discipline)

    if session["conditions"]:
        session["base_query1A"] += " WHERE " + " AND ".join(session["conditions"]) + " GROUP BY discipline; "
        session["base_query2"] += " WHERE " + " AND ".join(session["conditions"])
        session["base_query3"] += " WHERE " + " AND ".join(session["conditions"])
        session["base_query4"] += " WHERE " + " AND ".join(session["conditions"])
        session["base_query5"] += " WHERE " + " AND ".join(session["conditions"])
        session['float_days_query'] += " WHERE " + " AND ".join(session["conditions"])
        session['pre_dates_query'] += " WHERE " + " AND ".join(session["conditions"])
        session['post_dates_query'] += " WHERE " + " AND ".join(session["conditions"])
        session['average_time'] +=" WHERE " + " AND ".join(session["conditions"])   
        session['status_'] +=" WHERE " + " AND ".join(session["conditions"])
        session['total_pr_no_query']+=" WHERE " + " AND ".join(session["conditions"])
        session['process_ppi_query_1']+= " AND "+ " AND ".join(session["conditions"])
        session['process_ppi_query_2']+= " AND "+ " AND ".join(session["conditions"]) 
        session['process_ppi_query_3']+= " AND "+ " AND ".join(session["conditions"]) 
        session['process_ppi_query_4']+= " AND "+ " AND ".join(session["conditions"]) 
        session['process_ppi_query_5']+= " AND "+ " AND ".join(session["conditions"])    
        
        
        session['Mechanical_ppi_query_1']+=" AND "+ " AND ".join(session["conditions"])
        session['Mechanical_ppi_query_2']+=" AND "+ " AND ".join(session["conditions"]) 
        session['Mechanical_ppi_query_3']+=" AND "+ " AND ".join(session["conditions"]) 
        session['Mechanical_ppi_query_4']+=" AND "+ " AND ".join(session["conditions"]) 
        session['Mechanical_ppi_query_5']+=" AND "+ " AND ".join(session["conditions"])   

        session['Electrical_ppi_query_1']+= " AND "+ " AND ".join(session["conditions"])
        session['Electrical_ppi_query_2']+= " AND "+ " AND ".join(session["conditions"]) 
        session['Electrical_ppi_query_3']+= " AND "+ " AND ".join(session["conditions"]) 
        session['Electrical_ppi_query_4']+= " AND "+ " AND ".join(session["conditions"]) 
        session['Electrical_ppi_query_5']+= " AND "+ " AND ".join(session["conditions"])   
        
        session['Instrumentation_ppi_query_1']+= " AND " + " AND ".join(session["conditions"])
        session['Instrumentation_ppi_query_2']+= " AND " + " AND ".join(session["conditions"]) 
        session['Instrumentation_ppi_query_3']+= " AND " + " AND ".join(session["conditions"]) 
        session['Instrumentation_ppi_query_4']+= " AND " + " AND ".join(session["conditions"]) 
        session['Instrumentation_ppi_query_5']+= " AND " + " AND ".join(session["conditions"])   
        
    else:
        
        session["base_query1A"] += " GROUP BY discipline;"
         
    disciplines = ['Process', 'Mech', 'Electrical', 'Instru', 'Other']
    stages = ['not_yet_started', 'pre_order_stage', 'preorder_completed', 'post_order_stage', 'completed']


    for disciplines in disciplines:
        for stage in stages:
            session[f'{stage}_{disciplines.lower()}'] = 0
        
    cursor.execute(session["base_query1"] + session["base_query1A"], tuple(session["parameters"]))
    session["results"] = cursor.fetchall()
    cursor.execute(session['status_'], tuple(session["parameters"]))
    session["status_elapsed"] = cursor.fetchall()
    cursor.execute(session['average_time'], tuple(session["parameters"]))
    session["average_time_days"] = cursor.fetchall()
    session['days_lapsed']=elapsed_time(session["average_time_days"],session['status_elapsed'])
    cursor.execute(session["base_query2"], tuple(session["parameters"]))
   

    session['hotcolumns'] = [col[0] for col in cursor.description]
    session['hottemp'] = cursor.fetchall()
    for i, row in enumerate(session['hottemp']):
        row = list(row)
        row.pop(7)
        row.insert(7, session['days_lapsed'][i])
        session['hottemp'][i] = row 

    session['hotdata'] = [dict(zip(session['hotcolumns'], row)) for row in session['hottemp']]
    # print("q",session['process_ppi_query_1'])
    cursor.execute(session['total_pr_no_query'], tuple(session["parameters"]))
    # session['total_pr_no']= cursor.fetchone()[0]
    
    
    cursor.execute(session['process_ppi_query_1'], tuple(session["parameters"]))
    session['process_ppi_result_1']= cursor.fetchone()[0]
    print('ppi',session['process_ppi_result_1'])
    cursor.execute(session['process_ppi_query_2'], tuple(session["parameters"]))
    
    session['process_ppi_result_2']= cursor.fetchone()[0]
    cursor.execute(session['process_ppi_query_3'], tuple(session["parameters"]))

    session['process_ppi_result_3']= cursor.fetchone()[0]
    cursor.execute(session['process_ppi_query_4'], tuple(session["parameters"]))

    session['process_ppi_result_4']= cursor.fetchone()[0]
    cursor.execute(session['process_ppi_query_5'], tuple(session["parameters"]))

    session['process_ppi_result_5']= cursor.fetchone()[0]
    
    
    cursor.execute(session['Mechanical_ppi_query_1'], tuple(session["parameters"]))
    session['Mechanical_ppi_result_1']= cursor.fetchone()[0]
    print('ppi',session['Mechanical_ppi_result_1'])
    cursor.execute(session['Mechanical_ppi_query_2'], tuple(session["parameters"]))
    
    session['Mechanical_ppi_result_2']= cursor.fetchone()[0]
    cursor.execute(session['Mechanical_ppi_query_3'], tuple(session["parameters"]))

    session['Mechanical_ppi_result_3']= cursor.fetchone()[0]
    cursor.execute(session['Mechanical_ppi_query_4'], tuple(session["parameters"]))

    session['Mechanical_ppi_result_4']= cursor.fetchone()[0]
    cursor.execute(session['Mechanical_ppi_query_5'], tuple(session["parameters"]))

    session['Mechanical_ppi_result_5']= cursor.fetchone()[0]
    
    
    cursor.execute(session['Electrical_ppi_query_1'], tuple(session["parameters"]))
    session['Electrical_ppi_result_1']= cursor.fetchone()[0]
    print('ppi',session['Electrical_ppi_result_1'])
    cursor.execute(session['Electrical_ppi_query_2'], tuple(session["parameters"]))
    
    session['Electrical_ppi_result_2']= cursor.fetchone()[0]
    cursor.execute(session['Electrical_ppi_query_3'], tuple(session["parameters"]))

    session['Electrical_ppi_result_3']= cursor.fetchone()[0]
    cursor.execute(session['Electrical_ppi_query_4'], tuple(session["parameters"]))

    session['Electrical_ppi_result_4']= cursor.fetchone()[0]
    cursor.execute(session['Electrical_ppi_query_5'], tuple(session["parameters"]))

    session['Electrical_ppi_result_5']= cursor.fetchone()[0]
    
    
    cursor.execute(session['Instrumentation_ppi_query_1'], tuple(session["parameters"]))
    session['Instrumentation_ppi_result_1']= cursor.fetchone()[0]
    print('ppi',session['Instrumentation_ppi_result_1'])
    cursor.execute(session['Instrumentation_ppi_query_2'], tuple(session["parameters"]))
    
    session['Instrumentation_ppi_result_2']= cursor.fetchone()[0]
    cursor.execute(session['Instrumentation_ppi_query_3'], tuple(session["parameters"]))

    session['Instrumentation_ppi_result_3']= cursor.fetchone()[0]
    cursor.execute(session['Instrumentation_ppi_query_4'], tuple(session["parameters"]))

    session['Instrumentation_ppi_result_4']= cursor.fetchone()[0]
    cursor.execute(session['Instrumentation_ppi_query_5'], tuple(session["parameters"]))

    session['Instrumentation_ppi_result_5']= cursor.fetchone()[0]
    
  
    cursor.execute(session["base_query3"], tuple(session["parameters"]))
    session["kpi"] = cursor.fetchall()
    cursor.execute(session["base_query4"], tuple(session["parameters"]))
    session["kpi2"] = cursor.fetchall()
    cursor.execute(session["base_query5"], tuple(session["parameters"]))
    session["kpi2_sub"] = cursor.fetchall()
    cursor.execute(session['float_days_query'], tuple(session["parameters"]))
    session["float_dates"] = cursor.fetchall()
    cursor.execute(session['pre_dates_query'], tuple(session["parameters"]))
    session["pre_dates"] = cursor.fetchall()
    cursor.execute(session['post_dates_query'], tuple(session["parameters"]))
    session["post_dates"] = cursor.fetchall()
    
    calculate_average_days(session["average_time_days"] )
    
    
    
    for row in session["results"]:
        disciplines,  session['CompletedCount'], session['PostOrderStageCount'],session['PreOrderCompletedCount'],session['PreOrderStageCount'],session['NotYetStartedCount']= row
        if disciplines == "Process":
            session['not_yet_started_process'] = session['NotYetStartedCount']
            session['pre_order_stage_process'] = session['PreOrderStageCount']
            session['preorder_completed_process'] = session['PreOrderCompletedCount']
            session['post_order_stage_process'] = session['PostOrderStageCount']
            session['completed_process'] = session['CompletedCount']
        elif disciplines == "Mechanical":
            session['not_yet_started_mech'] = session['NotYetStartedCount']
            session['pre_order_stage_mech'] = session['PreOrderStageCount']
            session['preorder_completed_mech'] = session['PreOrderCompletedCount']
            session['post_order_stage_mech'] = session['PostOrderStageCount']
            session['completed_mech'] = session['CompletedCount']
        elif disciplines == "Electrical":
            session['not_yet_started_electrical'] = session['NotYetStartedCount']
            session['pre_order_stage_electrical'] = session['PreOrderStageCount']
            session['preorder_completed_electrical'] = session['PreOrderCompletedCount']
            session['post_order_stage_electrical'] = session['PostOrderStageCount']
            session['completed_electrical'] = session['CompletedCount']
        elif disciplines == "Instrumentation":
            session['not_yet_started_instru'] = session['NotYetStartedCount']
            session['pre_order_stage_instru'] = session['PreOrderStageCount']
            session['preorder_completed_instru'] = session['PreOrderCompletedCount']
            session['post_order_stage_instru'] = session['PostOrderStageCount']
            session['completed_instru'] = session['CompletedCount']
        elif disciplines == "other":
            session['not_yet_started_other'] = session['NotYetStartedCount']
            session['pre_order_stage_other'] = session['PreOrderStageCount']
            session['preorder_completed_other'] = session['PreOrderCompletedCount']
            session['post_order_stage_other'] = session['PostOrderStageCount']
            session['completed_other'] = session['CompletedCount']

    session['data'] = {
        'not_yet_started_process': session['not_yet_started_process'],
        'pre_order_stage_process': session['pre_order_stage_process'],
        'preorder_completed_process': session['preorder_completed_process'],
        'post_order_stage_process': session['post_order_stage_process'],
        'completed_process': session['completed_process'],

        'not_yet_started_mech': session['not_yet_started_mech'],
        'pre_order_stage_mech': session['pre_order_stage_mech'],
        'preorder_completed_mech': session['preorder_completed_mech'],
        'post_order_stage_mech': session['post_order_stage_mech'],
        'completed_mech': session['completed_mech'],

        'not_yet_started_electrical': session['not_yet_started_electrical'],
        'pre_order_stage_electrical': session['pre_order_stage_electrical'],
        'preorder_completed_electrical': session['preorder_completed_electrical'],
        'post_order_stage_electrical': session['post_order_stage_electrical'],
        'completed_electrical': session['completed_electrical'],

        'not_yet_started_instru': session['not_yet_started_instru'],
        'pre_order_stage_instru': session['pre_order_stage_instru'],
        'preorder_completed_instru': session['preorder_completed_instru'],
        'post_order_stage_instru': session['post_order_stage_instru'],
        'completed_instru': session['completed_instru'],

        'not_yet_started_other': session['not_yet_started_other'],
        'pre_order_stage_other': session['pre_order_stage_other'],
        'preorder_completed_other': session['preorder_completed_other'],
        'post_order_stage_other': session['post_order_stage_other'],
        'completed_other': session['completed_other'],

        "mech": session['not_yet_started_mech'] + session['pre_order_stage_mech'] + session['preorder_completed_mech'] + session['post_order_stage_mech'] + session['completed_mech'],
        "process": session['not_yet_started_process'] + session['pre_order_stage_process'] + session['preorder_completed_process'] + session['post_order_stage_process'] + session['completed_process'],
        "electrical": session['not_yet_started_electrical'] + session['pre_order_stage_electrical'] + session['preorder_completed_electrical'] + session['post_order_stage_electrical'] + session['completed_electrical'],
        "instru": session['not_yet_started_instru'] + session['pre_order_stage_instru'] + session['preorder_completed_instru'] + session['post_order_stage_instru'] + session['completed_instru'],
        "other": session['not_yet_started_other'] + session['pre_order_stage_other'] + session['preorder_completed_other'] + session['post_order_stage_other'] + session['completed_other'],

        "not_yet_started": session['not_yet_started_process'] + session['not_yet_started_mech'] + session['not_yet_started_electrical'] + session['not_yet_started_instru'] + session['not_yet_started_other'],
        "pre_order_stage": session['pre_order_stage_process'] + session['pre_order_stage_mech'] + session['pre_order_stage_electrical'] + session['pre_order_stage_instru'] + session['pre_order_stage_other'],
        "preorder_completed": session['preorder_completed_process'] + session['preorder_completed_mech'] + session['preorder_completed_electrical'] + session['preorder_completed_instru'] + session['preorder_completed_other'],
        "post_order_stage": session['post_order_stage_process'] + session['post_order_stage_mech'] + session['post_order_stage_electrical'] + session['post_order_stage_instru'] + session['post_order_stage_other'],
        "completed": session['completed_process'] + session['completed_mech'] + session['completed_electrical'] + session['completed_instru'] + session['completed_other'],
        "with_edrc_pre":session['average_days_EDRC_PRE'],
        "with_edrc_post":session['average_days_EDRC_POST'],
        "with_vendor_pre":session['average_days_VENDOR_PRE'],
        "with_vendor_post":session['average_days_VENDOR_POST']
        
    }
    
    
    
    for row in session["kpi"]:
        PendingWithVendorPRE, PendingWithEDRCPRE, Cleared =row #, 
        session['PendingWithVendorPRE'] = PendingWithVendorPRE
        session['PendingWithEDRCPRE'] = PendingWithEDRCPRE
        session['Cleared'] = Cleared
        session['total_pr']=(session['data']["process"]+session['data']["mech"]+session['data']["electrical"]+session['data']["instru"]+session['data']["other"])-(session['not_yet_started_mech']+session['not_yet_started_process']+session['not_yet_started_instru']+session['not_yet_started_other']+session['not_yet_started_electrical'])
        
    for row in session["kpi2"]:
        PendingWithVendorPOST, PendingWithEDRCPOST, Completed=row
        session['PendingWithVendorPOST'] = PendingWithVendorPOST
        session['PendingWithEDRCPOST'] = PendingWithEDRCPOST
        session['Completed'] = Completed
        session['poissued'] = session['data']["post_order_stage"]+session['data']["completed"]
        # print(session['data']["completed"])
        
        
        
    for row in session["kpi2_sub"]:
        total_pr_cleared,ClearedPO=row
        session['ClearedPO']=ClearedPO
        session['total_pr_cleared']=total_pr_cleared
    
    session['floated']=cycle_time(session["float_dates"])
    session['pre_o']=cycle_time(session["pre_dates"])
    session['post_o']=cycle_time(session["post_dates"])
    
   
    session['total_pr_no_process']= session['not_yet_started_process'] + session['pre_order_stage_process'] + session['preorder_completed_process'] + session['post_order_stage_process'] + session['completed_process']  
    session['total_pr_no_mech']=session['not_yet_started_mech'] + session['pre_order_stage_mech'] + session['preorder_completed_mech'] + session['post_order_stage_mech'] + session['completed_mech'] 
    session['total_pr_no_elec']= session['not_yet_started_electrical'] + session['pre_order_stage_electrical'] + session['preorder_completed_electrical'] + session['post_order_stage_electrical'] + session['completed_electrical']     
    session['total_pr_no_instru']= session['not_yet_started_instru'] + session['pre_order_stage_instru'] + session['preorder_completed_instru'] + session['post_order_stage_instru'] + session['completed_instru']   
    
    session['total_pr_no']=session['total_pr_no_process']+session['total_pr_no_mech']+session['total_pr_no_elec']+session['total_pr_no_instru'] 
    
    
    print('pr',session['total_pr_no'])
    session['process_ppi']=round(((int(session['process_ppi_result_1'])/int(session['total_pr_no_process']))*0.15+(int(session['process_ppi_result_2'])/int(session['total_pr_no_process']))*0.3+(int(session['process_ppi_result_3'])/int(session['total_pr_no_process']))*0.15+(int(session['process_ppi_result_4'])/int(session['total_pr_no_process']))*0.15+(int(session['process_ppi_result_5'])/int(session['total_pr_no_process']))*0.25)*100,2)
    print('Process PPI Result',session['process_ppi'])
   
    session['Mechanical_ppi']=round(((int(session['Mechanical_ppi_result_1'])/int(session['total_pr_no_mech']))*0.15+(int(session['Mechanical_ppi_result_2'])/int(session['total_pr_no_mech']))*0.3+(int(session['Mechanical_ppi_result_3'])/int(session['total_pr_no_mech']))*0.15+(int(session['Mechanical_ppi_result_4'])/int(session['total_pr_no_mech']))*0.15+(int(session['Mechanical_ppi_result_5'])/int(session['total_pr_no_mech']))*0.25)*100,2)
    print('Mechanical PPI Result',session['Mechanical_ppi'])
   
    session['Electrical_ppi']=round(((int(session['Electrical_ppi_result_1'])/int(session['total_pr_no_elec']))*0.15+(int(session['Electrical_ppi_result_2'])/int(session['total_pr_no_elec']))*0.3+(int(session['Electrical_ppi_result_3'])/int(session['total_pr_no_elec']))*0.15+(int(session['Electrical_ppi_result_4'])/int(session['total_pr_no_elec']))*0.15+(int(session['Electrical_ppi_result_5'])/int(session['total_pr_no_elec']))*0.25)*100,2)
    print('Electrical PPI Result',session['Electrical_ppi'])
    session['Instrumentation_ppi']=round(((int(session['Instrumentation_ppi_result_1'])/int(session['total_pr_no_instru']))*0.15+(int(session['Instrumentation_ppi_result_2'])/int(session['total_pr_no_instru']))*0.3+(int(session['Instrumentation_ppi_result_3'])/int(session['total_pr_no_instru']))*0.15+(int(session['Instrumentation_ppi_result_4'])/int(session['total_pr_no_instru']))*0.15+(int(session['Instrumentation_ppi_result_5'])/int(session['total_pr_no_instru']))*0.25)*100,2)
    print('Instrumentation PPI Result',session['Instrumentation_ppi'])
    
    session['total_ppi']=round((session['Mechanical_ppi']*40+session['process_ppi']*25+session['Electrical_ppi']*20+session['Instrumentation_ppi']*5)/100,2)
    conn.close()
    return (session['data'])



@db_bp.route('/suggest', methods=['POST'])
def suggest():
    SUGGEST_DOCSTRING
    try:
        session['search_query'] = request.form.get('search_query')

        job_code_suggestions = Project.query.filter(Project.job_code.ilike(f'%{session["search_query"]}%')).limit(10).all()
        project_name_suggestions = Project.query.filter(Project.project_name.ilike(f'%{session["search_query"]}%')).limit(5).all()

        
        suggestions = [
            f'{project.job_code} - {project.project_name}' for project in job_code_suggestions
        ] + [
            f'{project.job_code} - {project.project_name}' for project in project_name_suggestions
        ]

        return jsonify(suggestions)

    except Exception as e:
        return jsonify({'error': str(e)})



@db_bp.route('/api/get_kpi_data')
def kpidata():
    KPIDATA_DOCSTRING
    if session['floated'] is None:
            session['floated']="-"
    if session['pre_o'] is None:
            session['pre_o']="-"
    if session['post_o'] is None:
            session['post_o']="-"
            
    session['kpi_data'] = {
        "PendingWithVendorPRE":session['PendingWithVendorPRE'],
        "PendingWithEDRCPRE":session['PendingWithEDRCPRE'],
        'Cleared':session['Cleared'],
        "prissued":session['total_pr'],
        "prcleared":session['total_pr_cleared'],
        'PendingWithVendorPOST':session['PendingWithVendorPOST'],
        'PendingWithEDRCPOST':session['PendingWithEDRCPOST'],
        'Completed':session['Completed'],
        'poissued':session['poissued'],
        'ClearedPO':session['ClearedPO'],
        'floated':session['floated'],
        'pre_o':session['pre_o'],
        'post_o':session['post_o'],
        
        
    }
   
    return jsonify(data=session['kpi_data'])




@db_bp.route('/saveProject', methods=['POST'])
def save_project():
    SAVE_PROJECT_DOCSTRING
    # try:
    session['ic'] = request.form.get("ic")
    session['sbg'] = request.form.get("SBG")
    session['bu'] = request.form.get("BU")
    session['emanage'] = request.form.get("Emanage")
    session['plead'] = request.form.get("Plead")
    session['mlead'] = request.form.get("Mlead")
    session['elead'] = request.form.get("Elead")
    session['ilead'] = request.form.get("Ilead")
    session['others'] = request.form.get("others")
    session['project'] = request.form.get("Project")
    session['client'] = request.form.get("Cname")
    session['Jobcode'] = request.form.get("Jobcode")
    session['scmlead'] = request.form.get("scmlead") 
    session['process_date'] = request.form.get("process_date")
    session['mech_date'] = request.form.get("mech_date")
    session['elec_date'] = request.form.get("elec_date")
    session['instru_date'] = request.form.get("instru_date")
    session['other_date'] = request.form.get("other_date")
    

    if_jobcode = Project.query.filter_by(job_code=session['Jobcode']).first()
    print(if_jobcode)
    if if_jobcode:

            return f"""
        <script>
            alert('Project Already Registered');
            window.location.href = '/lnt/newproject';
        </script>
        """ 
    else:

        
        new_project = Project(
            ic=session['ic'],
            sbg=session['sbg'],
            bu=session['bu'],
            emanage=session['emanage'],
            plead=session['plead'],
            mlead=session['mlead'],
            elead=session['elead'],
            ilead=session['ilead'],
            others=session['others'],
            project_name=session['project'],
            client_name=session['client'],
            job_code=session['Jobcode'],
            scmlead=session['scmlead'],
            process_date=session['process_date'],
            mech_date=session['mech_date'],
            elec_date=session['elec_date'],
            instru_date=session['instru_date'],
            other_date=session['other_date'],
            
        )

        db.session.add(new_project)
        # db.session.add(new_pr_status)
        
        db.session.commit()

        # Return a success response with a script
        return f"""
        <script>
            alert('Project saved successfully!');
            window.location.href = '/lnt/newproject';
        </script>
        """

@db_bp.route('/getProjectDetails', methods=['GET'])
def get_project_details():
    GET_PROJECT_DETAILS_DOCSTRING
    try:
        session['job_code'] = request.args.get('job_code')
        session['job_code'] =session['job_code'][:8] 
        
        
        # Example SQLAlchemy query to get project details
        session['project'] = Project.query.filter_by(job_code=session['job_code']).first()
        print(session['project'])

        if session['project']:
            session['result'] = {
                'ic': session['project'].ic,
                'sbg': session['project'].sbg,
                'bu': session['project'].bu,
                'emanage': session['project'].emanage,
                'plead': session['project'].plead,
                'mlead': session['project'].mlead,
                'elead': session['project'].elead,
                'ilead': session['project'].ilead,
                'scmlead': session['project'].scmlead,
                'process_date': session['project'].process_date,
                'mech_date': session['project'].mech_date,
                'elec_date': session['project'].elec_date,
                'instru_date': session['project'].instru_date,
                'other_date': session['project'].other_date,
                
                'others': session['project'].others,
                'project_name': session['project'].project_name,
                'client_name': session['project'].client_name,
                'job_code': session['project'].job_code,
            }
            return jsonify(session['result'])
        else:
            return jsonify({'error': 'Project not found'})

    except Exception as e:
        return jsonify({'error': str(e)})
    

@db_bp.route('/updateProject', methods=['POST'])
def update_project():
        UPDATE_PROJECT_DOCSTRING
        # Get the updated data from the request
        updated_data = {
            'emanage': request.form.get('emanage'),
            'plead': request.form.get('plead'),
            'mlead': request.form.get('mlead'),
            'elead': request.form.get('elead'),
            'ilead': request.form.get('ilead'),
            'others': request.form.get('others'),
            'scmlead': request.form.get('scmlead'),
            'process_date':request.form.get('process_date'),
            'mech_date':request.form.get('mech_date'),
            'elec_date':request.form.get('elec_date'),
            'instru_date':request.form.get('instru_date'),
            'other_date':request.form.get('other_date')
            
        }
        print('ud       =>>>>>',updated_data)
        
        job_code = request.form.get('Jc')
        job_code = job_code[:8]
        print(job_code)

        project = Project.query.filter_by(job_code=job_code).first()
        print(project)
        if project:
            project.emanage = updated_data['emanage']
            project.plead = updated_data['plead']
            project.mlead = updated_data['mlead']
            project.elead = updated_data['elead']
            project.ilead = updated_data['ilead']
            project.scmlead = updated_data['scmlead']
            project.process_date=updated_data['process_date']
            project.mech_date=updated_data['mech_date']
            project.elec_date=updated_data['elec_date']
            project.instru_date=updated_data['instru_date']
            project.other_date=updated_data['other_date']
            project.others = updated_data['others']
            db.session.commit()
            return f"""
                <script>
                    alert('Project Updated Successfully");
                    window.location.href = '/lnt/manageproject';
                </script>
                """
        else:
            return jsonify({'success': False, 'message': 'Project not found'})
    



@db_bp.route('/api/getData')
def get_data():
    try:
        # Execute an SQL query to retrieve data
        conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')  
        cursor = conn.cursor()
        cursor.execute('SELECT EMI_ITEMS,Vendor_Name,Person,Contact_No,E_Mail,Location,Designation,Scope_of_Work,Address,Type_Of_Supplier FROM vendorDataBank')  
        session['columns'] = [col[0] for col in cursor.description]
        session['data'] = [dict(zip(session['columns'], row)) for row in cursor.fetchall()]
        # Close the database connection
        conn.close()

        return jsonify({'columns': session['columns'], 'data': session['data']})

    except Exception as e:
        return jsonify({'error': str(e)})
    

    
@db_bp.route('/api/getprStatus')  
def get_statusdata():
    GET_STATUSDATA_DOCSTRING
    try:
        return jsonify({'columns': session['hotcolumns'], 'data': session['hotdata'],"days":session['days_lapsed']})

    except Exception as e:
        return jsonify({'error': str(e)})

@db_bp.route('/api/getPrdata', methods=['POST'])
def details():
    DETAILS_DOCSTRING
    session['email'] = get_email()
    
    session['selected_job_code'] = request.form.get('selected_job_code')
    session['selected_job_code'] = session['selected_job_code'][:8]

    conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')
    cursor = conn.cursor()

    cursor.execute('SELECT emanage,plead,mlead,elead,ilead,scmlead,others FROM project WHERE job_code=?', (session['selected_job_code'],))
    session['admins'] = cursor.fetchall()[0]

    session['other_emails'] = session['admins'][6].split(';')
    session['scm_emails'] = session['admins'][5].split(';') if session['admins'][5]!=None else None
    

    # Define discipline to email mapping
    session['discipline_emails'] = {
        'Process': set(session['admins'][1].split(';')),
        'Mechanical': set(session['admins'][2].split(';')),
        'Electrical': set(session['admins'][3].split(';')),
        'Instrumentation': set(session['admins'][4].split(';')),
        'others': set(session['admins'][6].split(';'))
    }
    print(session['email'])
    
    # Check if the user's email is in emanage emails
    if session['email'].lower() in [email.lower() for email in session['admins'][0].split(';')] or session['email'].lower() in [email.lower() for email in master_emails]:
        print("User is Manager/Master, fetching all data")
        cursor.execute('''SELECT pr_issue_date_final,offer_recv_date,
                            comments_sent_pre,offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                            offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,
                            final_vendor_doc_pre,sent_to_scm,po_issued_date,
                            post_doc_rcv_dt,comments_sent_post,
                            final_vendor_doc_post,mfc_issued_date FROM pr_data WHERE job_code=?''', (session['selected_job_code'],))
        session["average_time_days_update"] = cursor.fetchall()
        cursor.execute('''SELECT status,item_code FROM pr_data WHERE job_code=?''',(session['selected_job_code'],))
        session['status_update']=cursor.fetchall()
        session['days_lapsed_update']=elapsed_time(session["average_time_days_update"],session['status_update'])
        cursor.execute('''SELECT project_name,discipline,equipment,item_code,item_vendor_code,vendor_name,status,elapsed_days,remarks,vendor_ph,vendor_mail,pr_issue_date_pre,pr_issue_date_final,
                       offer_recv_date,comments_sent_pre,offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                            offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,tc_meeting_vendor_pre,final_vendor_doc_pre,
                            sent_to_scm,priority,po_issued_date,finalized_vendor,post_doc_rcv_dt,comments_sent_post,tc_meeting_vendor_post,final_vendor_doc_post,
                            mfc_issued_date,target_completion_date,  target_completion_date_po,target_completion_date_post,last_modified_by FROM pr_data WHERE job_code=?''', (session['selected_job_code'],))
        session['hotupdate']=cursor.fetchall()
        session['columns'] = [col[0] for col in cursor.description]
        for i, row in enumerate(session['hotupdate']):
            row = list(row)
            row.pop(7)
            row.insert(7, session['days_lapsed_update'][i])
            session['hotupdate'][i] = row 
        
        session['data'] = [dict(zip(session['columns'], row)) for row in session['hotupdate']]
        conn.close()
        return jsonify({'columns': session['columns'], 'data': session['data'],"flag_po":0})
    elif  session['scm_emails']!= None and session['email'].lower() in [email.lower() for email in session['scm_emails']]:
        print("User is SCM, fetching all data")
        cursor.execute('''SELECT project_name,discipline,equipment,item_code,item_vendor_code,vendor_name,status,elapsed_days,remarks,vendor_ph,vendor_mail,pr_issue_date_pre,pr_issue_date_final,
                       offer_recv_date,comments_sent_pre,offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                            offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,tc_meeting_vendor_pre,final_vendor_doc_pre,
                            sent_to_scm,priority,po_issued_date,finalized_vendor,post_doc_rcv_dt,comments_sent_post,tc_meeting_vendor_post,final_vendor_doc_post,
                            mfc_issued_date,target_completion_date,  target_completion_date_po,target_completion_date_post, last_modified_by FROM pr_data WHERE job_code=?''', (session['selected_job_code'],))
        session['columns'] = [col[0] for col in cursor.description]
        session['data'] = [dict(zip(session['columns'], row)) for row in cursor.fetchall()]
        conn.close()
        return jsonify({'columns': session['columns'], 'data': session['data'],"flag_po":1})
    else:
        session['user_discipline'] = None
        for discipline, emails in session['discipline_emails'].items():
            if session['email'].lower() in [email.lower() for email in emails]:
                session['user_discipline'] = discipline
                break

        if session['user_discipline']:
            # User belongs to a specific discipline, fetch data accordingly
            print("Discipline:", session['user_discipline'])
            cursor.execute('''SELECT pr_issue_date_final,offer_recv_date,
                            comments_sent_pre,offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                            offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,
                            final_vendor_doc_pre,sent_to_scm,po_issued_date,
                            post_doc_rcv_dt,comments_sent_post,
                            final_vendor_doc_post,mfc_issued_date FROM pr_data WHERE job_code=? AND discipline=?''', (session['selected_job_code'],session['user_discipline'],))
            session["average_time_days_update"] = cursor.fetchall()
            cursor.execute('''SELECT status,item_code FROM pr_data WHERE job_code=? AND discipline=?''',(session['selected_job_code'],session['user_discipline'],))
            session['status_update']=cursor.fetchall()
            session['days_lapsed_update']=elapsed_time(session["average_time_days_update"],session['status_update'])
            cursor.execute('''SELECT project_name,discipline,equipment,item_code,item_vendor_code,vendor_name,status,elapsed_days,remarks,vendor_ph,vendor_mail,pr_issue_date_pre,pr_issue_date_final,offer_recv_date,comments_sent_pre,offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                            offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,tc_meeting_vendor_pre,final_vendor_doc_pre,sent_to_scm,priority,po_issued_date,finalized_vendor,post_doc_rcv_dt,comments_sent_post,tc_meeting_vendor_post,final_vendor_doc_post,
                            mfc_issued_date,target_completion_date,  target_completion_date_po,target_completion_date_post, last_modified_by FROM pr_data WHERE job_code=? AND discipline=?''', (session['selected_job_code'], session['user_discipline']))
            session['hotupdate']=cursor.fetchall()
            session['columns'] = [col[0] for col in cursor.description]
            for i, row in enumerate(session['hotupdate']):
                row = list(row)
                row.pop(7)
                row.insert(7, session['days_lapsed_update'][i])
                session['hotupdate'][i] = row 
            
            session['data'] = [dict(zip(session['columns'], row)) for row in session['hotupdate']]
            conn.close()
            return jsonify({'columns': session['columns'], 'data': session['data'],"flag_po":0})
        else:
            return render_template('nopermission.html'), 403


@db_bp.route('/delete_vendor', methods=['POST'])
def delete_pr_entry():
    session['email'] = get_email()
    print(session['email'],"Delete")
    session['vendor_number'] = request.json.get("vendor_number")
    form_t = time.strftime("-%d.%m.%Y-%H:%M:%S") 
    session['email_final']=session['email']+form_t
    if not session['vendor_number']:
        return jsonify({'error': 'Vendor number is required'}), 400

    conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')
    cursor = conn.cursor()
    cursor.execute("""
  INSERT INTO pr_data_backup ( ic, sbg, bu,job_code,project_name, discipline, equipment, item_code, item_vendor_code, 
                        vendor_name, status, remarks,pr_issue_date_pre,pr_issue_date_final,
                       offer_recv_date,comments_sent_pre,offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                            offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,tc_meeting_vendor_pre,final_vendor_doc_pre,
                            sent_to_scm,priority,po_issued_date,finalized_vendor,post_doc_rcv_dt,comments_sent_post,tc_meeting_vendor_post,final_vendor_doc_post,
                            mfc_issued_date,target_completion_date,  target_completion_date_po,target_completion_date_post,last_modified_by) 
                        
SELECT  ic, sbg, bu,job_code,project_name, discipline, equipment, item_code, item_vendor_code, 
                        vendor_name, status, remarks, pr_issue_date_pre,pr_issue_date_final,
                       offer_recv_date,comments_sent_pre,offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                            offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,tc_meeting_vendor_pre,final_vendor_doc_pre,
                            sent_to_scm,priority,po_issued_date,finalized_vendor,post_doc_rcv_dt,comments_sent_post,tc_meeting_vendor_post,final_vendor_doc_post,
                            mfc_issued_date,target_completion_date,  target_completion_date_po,target_completion_date_post, last_modified_by FROM pr_data WHERE item_vendor_code = ?""", (session['vendor_number'],))
    # cursor.execute('SELECT * FROM pr_data_backup WHERE item_vendor_code = ?', (session['vendor_number'],))
    # result = cursor.fetchone()

    # # If data is found, insert it into pr_data_backup
    # if result:
    cursor.execute('UPDATE pr_data_backup SET removed_by=? WHERE item_vendor_code=? ' ,(session['email_final'], session['vendor_number']))
    
    cursor.execute('DELETE FROM pr_data WHERE item_vendor_code = ?', (session['vendor_number'],))
    
    conn.commit()
    if cursor.rowcount == 0:
        return jsonify({'error': 'PR not found'}), 404

    return jsonify({'message': 'Deleted successfully!'}), 200


@db_bp.route('/api/updatePrdata', methods=['POST'])
def update_pr_data():
    UPDATE_PR_DATA_DOCSTRING 
    session['email'] = get_email()
    session['request_data'] = request.get_json()
    
    session['selected_job_code'] = session['request_data'].get('selected_job_code')  # Get the selected job code
    session['selected_job_code'] = session['selected_job_code'][:8]
    session['updated_data'] = session['request_data'].get('updated_data')
    print(session['email'])
    session['statuscheckdf'] = pd.DataFrame(session['updated_data'])
    session['statuses_to_check'] = [
        "Awaiting post-offer from vendor",
        "Post-order pending with EDRC",
        "Post-order comments issued",
        "Post-order technically cleared",
        "Approved by Client"
    ]
    session['other_statuses_to_check'] = ["Regret offer", "Sent to BU Operations","Pre-order technically rejected"]
    session['errors'] = []
    session['pr_code']=[]
    session['pr_v_code']=[]
    # Iterate over each row in the DataFrame
    for index, row in session['statuscheckdf'].iterrows():
        if row[6] in session['statuses_to_check']:
            print(row[4])
            session['pr_code'].append(row[3])
            session['pr_v_code'].append(row[4])    
    for index, row in session['statuscheckdf'].iterrows(): 
        if  row[3] in  session['pr_code'] and row[4] not in session['pr_v_code']:    
            if row[6] not in session['other_statuses_to_check']:
                session['r_num']=index+1
                session['errors'].append(f"PR Vendor code- {row[3]} has incorrect status in row- {session['r_num']}")
                    
    if session['errors']:
        for error in session['errors']:
             return jsonify({'success': False, 'error':error }),403
    else:
        print("No session['errors'] found.")

    special_characters = ['!', '"', '#', '$', '%', "'", '(', ')', '*', '+', ',', '-', '.', '/', ':', ';', '<', '=', '>', '?', '@', '[', '\\', ']', '^', '_', '`', '{', '|', '}', '~']

    conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')
    cursor = conn.cursor()
    session["rownum"]=0
    for row in session['updated_data']:
        session["rownum"]+=1  
        if not row[1] or row[1].strip() == '' or row[1] is None :
            print("Failed")
            return jsonify({'success': False, 'error': 'Please fill all the mandatory fields- Discipline '}),403
        if not row[2] or row[2].strip() == '' or row[2] is None :
            print("Failed")
            return jsonify({'success': False, 'error': 'Please fill all the mandatory fields- Equipment Name'}),403     
        if not row[3] or row[3].strip() == '' or row[3] is None :
            print("Failed")
            return jsonify({'success': False, 'error': 'Please fill all the mandatory fields- PR Code'}),403
        if not row[4] or row[4].strip() == '' or row[4] is None :
            print("Failed")
            return jsonify({'success': False, 'error': 'Please fill all the mandatory fields- PR Vendor Code'}),403
        else:
            stages = [
    "Enquiry yet to float",
    "Awaiting pre-offer from vendor",
    "Pre-order pending with EDRC",
    "Pre-order comments issued",
    "Pre-order technically cleared",
    "Pre-order technically rejected",
    "Regret offer",
    "Sent to BU Operations",
    "Awaiting post-offer from vendor",
    "Post-order pending with EDRC",
    "Post-order comments issued",
    "Post-order technically cleared",
    "Approved by Client"
]
            if row[6] not in stages[0:13]:   
                return jsonify({'success': False, 'error': 'Please choose the Correct status in row: '+ str(session["rownum"])}),403             
               
            if row[6] in stages[1:13]:               
                if row[5] is  None or row[12] is None or row[5].strip()=="" or row[12].strip()==""  :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory fields- Vendor Name & PR Enquiry Date in row: '+ str(session["rownum"])}),403
                if row[5]:
                    for char in row[5]:
                        if char in special_characters:
                            print("Failed")
                            return jsonify({'success': False, 'error': 'Do not enter Special Characters in Vendor Name. Remove: '+ char +' from: '+ row[5]}),403
            if row[6] in stages[2:6]:
                if row[13] is  None or row[13].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Offer recieved Date in row: '+ str(session["rownum"])}),403
            if row[6] in stages[3:5]:
                if row[14] is None or row[14].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Comments Sent from EDRC in row: '+ str(session["rownum"])}),403
            if row[6] == stages[4]:
                if row[23] is None or row[24] is None or row[23].strip()=="" or row[24].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory fields- Final Technical Clearance with vendor & Final Doc from Vendor in row: '+ str(session["rownum"])}),403
            if row[6] in stages[7:13]:
                if  row[13] is None or row[13].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Offer recieved Date in row: '+ str(session["rownum"])}),403
                if  row[14] is None or row[14].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Comments Sent from EDRC in row: '+ str(session["rownum"])}),403
                if  row[16] is not  None and row[16].strip()!="" :
                    if  row[15] is None or row[15].strip()=="" :
                        return jsonify({'success': False, 'error': 'Please fill the mandatory field- Offer recieved Date(Rev B) in row: '+ str(session["rownum"])}),403
                if  row[18] is not  None and row[18].strip()!="" :
                    if  row[17] is None or row[17].strip()=="" :
                        return jsonify({'success': False, 'error': 'Please fill the mandatory field- Offer recieved Date(Rev C) in row: '+ str(session["rownum"])}),403
                if  row[20] is not  None and row[20].strip()!="" :
                    if  row[19] is None or row[19].strip()=="" :
                        return jsonify({'success': False, 'error': 'Please fill the mandatory field- Offer recieved Date(Rev D) in row: '+ str(session["rownum"])}),403  
                if  row[22] is not  None and row[22].strip()!="" :
                    if  row[21] is None or row[21].strip()=="" :
                        return jsonify({'success': False, 'error': 'Please fill the mandatory field- Offer recieved Date(Rev E) in row: '+ str(session["rownum"])}),403  
                    
                if  row[23] is None or row[23].strip()=="" :
                    
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Final Technical Clearance with vendor in row: '+ str(session["rownum"])}),403
                if  row[24] is None or row[24].strip()=="":
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Final Document recieved from Vendor in row: '+ str(session["rownum"])}),403
                if  row[25] is None or row[25].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Sent to BU Operations in row: '+ str(session["rownum"])}),403
                if  row[26] is None or row[26].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Priority in row: '+ str(session["rownum"])}),403    
            if row[6] in stages[8:13]:
                if row[27] is None or row[28] is None or row[27].strip()=="" or row[28].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory fields- Finalized Vendor Name & PO Issued Date in row: '+ str(session["rownum"])}),403
            if row[6] in stages[9:13]:
                if row[29] is None or row[29].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Post order document recieved date in row: '+ str(session["rownum"])}),403
            if row[6] in stages[10:13]:
                if row[30] is None or row[30].strip()==""  :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Comments sent from EDRC date in row: '+ str(session["rownum"])}),403
            if row[6] in stages[11:13]:
                if  row[31] is None or row[31].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Final Technical Clearance meeting in row: '+ str(session["rownum"])}),403
            if row[6] in stages[11:13]:
                if row[32] is None or row[32].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- Final document recieved from Vendor in row: '+ str(session["rownum"])}),403
            if row[6]==stages[12]:
                if  row[33] is None or row[33].strip()=="" :
                    return jsonify({'success': False, 'error': 'Please fill the mandatory field- MFC Issued Date in row: '+ str(session["rownum"])}),403
            
            
            session['item_vendor_code'] = row[4]  
            cursor.execute("SELECT * FROM pr_data WHERE job_code=? AND item_vendor_code=?", (session['selected_job_code'], session['item_vendor_code']))
            session['existing_row'] = cursor.fetchone()
            
            if session['existing_row']:
                print("existing")
                cursor.execute("""
                    SELECT * FROM pr_data 
                    WHERE job_code=? AND item_vendor_code=?
                """, (session['selected_job_code'], session['item_vendor_code']))

                session['existing_row'] = cursor.fetchone()
                
                cursor.execute("""
                    SELECT project_name,vendor_name,finalized_vendor FROM pr_data 
                    WHERE job_code=? AND item_vendor_code=?
                """, (session['selected_job_code'], session['item_vendor_code']))
                
                session["vendor_metadata"] = cursor.fetchone()
                
                session['project_name'],session['name_of_vendor'],session['final_vendor']=session["vendor_metadata"]
                
                if  row[28]==session['name_of_vendor'] or  row[28] is None or row[28].strip()=="":
                    if row[28] is None:
                        row[28]=" "
                    
                    # session['item_code'] = row[3].replace(" ", "") if row[3] else None
                    # session['item_vendor_code'] = row[4].replace(" ", "") if row[4] else None
                    print("came")
                    cursor.execute("""
                        UPDATE pr_data 
                        SET project_name=?, discipline=?, equipment=?, item_code=?, item_vendor_code=?, 
                            vendor_name=?, status=?,elapsed_days=?, remarks=?,vendor_ph=?, vendor_mail=?, pr_issue_date_pre=?, pr_issue_date_final=?, 
                            offer_recv_date=?, comments_sent_pre=?,offer_recv_date_2=?,comments_sent_pre_2=?,offer_recv_date_3=?,comments_sent_pre_3=?,
                            offer_recv_date_4=?,comments_sent_pre_4=?,offer_recv_date_5=?,comments_sent_pre_5=?, tc_meeting_vendor_pre=?, 
                            final_vendor_doc_pre=?, sent_to_scm=?, priority=?, po_issued_date=?, 
                            finalized_vendor=?, post_doc_rcv_dt=?, comments_sent_post=?, 
                            tc_meeting_vendor_post=?, final_vendor_doc_post=?, mfc_issued_date=?, 
                            target_completion_date=?,target_completion_date_po=?,target_completion_date_post=?,last_modified_by=?
                        WHERE job_code=? AND item_vendor_code=?
                    """, (
                        session['project_name'], row[1], row[2],row[3], row[4],  row[5], row[6], row[7], 
                        row[8], row[9], row[10], row[11], row[12], row[13], row[14], 
                        row[15], row[16], row[17], row[18], row[19], row[20], row[21], row[22], row[23],row[24],
                        row[25],row[26],row[27], row[28], row[29], row[30], row[31],row[32],row[33],row[34],row[35],row[36],row[37],
                        session['selected_job_code'], session['item_vendor_code'],
                    ))

                    # Check if the row was updated
                    if cursor.rowcount > 0:
                        print("Row updated successfully")

                        # Retrieve the updated row
                        cursor.execute("""
                            SELECT * FROM pr_data 
                            WHERE job_code=? AND item_vendor_code=?
                        """, (session['selected_job_code'], session['item_vendor_code']))
                        form_t = time.strftime("-%d.%m.%Y-%H:%M:%S") 
                        session['modified_data']=session['email']+form_t
                        # print(session['modified_data'])
                        updated_row = cursor.fetchone()
                        # Compare the existing row with the updated row
                        if session['existing_row'] != updated_row:
                            cursor.execute("""
                                UPDATE pr_data 
                                SET last_modified_by=?
                                WHERE job_code=? AND item_vendor_code=?
                            """, (session['modified_data'], session['selected_job_code'], session['item_vendor_code']))

                        else:
                            print("No changes detected")

                    else:
                        print("Row not found or not changed")
                
                else:
                    return jsonify({'success': False, 'error': 'Please enter correct vendor name in Row: '+str(session["rownum"])+", Vendor: "+row[28]}),403



            else:
                cursor.execute("SELECT ic, sbg, bu,project_name FROM project WHERE job_code=?", (session['selected_job_code'],))
                session['project_metadata'] = cursor.fetchone()

                if not session['project_metadata']:
                    return jsonify({'success': False, 'error': 'Project metadata not found'})
                form_t = time.strftime("-%d.%m.%Y-%H:%M:%S") 
                session['modified_data']=session['email']+form_t
                session['ic'], session['sbg'], session['bu'] , session['project_name'] = session['project_metadata']
                print(session['project_name'])
                session['item_code'] = row[3].replace(" ", "") if row[3] else None
                session['item_vendor_code'] = row[4].replace(" ", "") if row[4] else None
                session['combined_data'] = (
                    session['selected_job_code'], session['ic'], session['sbg'], session['bu'],session['project_name'],
                    row[1],row[2], session['item_code'], session['item_vendor_code'], row[5], row[6], row[7], row[8], row[9],
                    row[10], row[11], row[12], row[13], row[14], row[15], row[16], row[17],
                    row[18], row[19], row[20], row[21], row[22], row[23],row[24],row[25],row[26] ,row[27], row[28], row[29], row[30], row[31],row[32],row[33],row[34],row[35],row[36],session['modified_data']
                )

                cursor.execute("""
                    INSERT INTO pr_data (job_code, ic, sbg, bu,project_name, discipline, equipment, item_code, item_vendor_code, 
                        vendor_name, status,elapsed_days, remarks,vendor_ph, vendor_mail,pr_issue_date_pre, pr_issue_date_final, 
                        offer_recv_date, comments_sent_pre, offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                            offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,tc_meeting_vendor_pre, 
                        final_vendor_doc_pre,sent_to_scm,priority,po_issued_date, finalized_vendor, post_doc_rcv_dt, 
                        comments_sent_post, tc_meeting_vendor_post, final_vendor_doc_post,mfc_issued_date,target_completion_date,  target_completion_date_po,target_completion_date_post,last_modified_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?, ?, ?, ?,?,?,?,?,?,?, ?, ?,?,?,?,?,?,?,?)
                """, (session['combined_data']))
  
   
    conn.commit()  # Commit the changes to the database
    conn.close()
    return jsonify({'success': True})



@db_bp.route('/dashboard')
def dashboard():
    DASHBOARD_DOCSTRING
    if "user" in session:
        userr = session["user"]
        print("dashboard",userr)
        session['data'] = chart('Discipline','Project','BU','sbg','IC')
        ic_options = fetch_dropdown_options('ic') 
        sbg = fetch_dropdown_options('sbg')
        bu = fetch_dropdown_options('bu')
        project = fetch_dropdown_options('project_name')
        discipline_filter = fetch_discipline_options('discipline')
        session['selected_ic'] = "ALL"
        session['selected_sbg'] ="ALL"
        session['selected_bu'] = "ALL"
        session['selected_project'] = "ALL"
        session['discipline_filter'] ="ALL"
        print(session['discipline_filter'])
        return render_template('dashboard.html',title='Management Console',ic_options=ic_options,bu=bu,sbg=sbg,project=project,discipline=discipline_filter,process_ppi=session['process_ppi'],mech_ppi=session['Mechanical_ppi'],elec_ppi=session['Electrical_ppi'],instru_ppi=session['Instrumentation_ppi'],total_ppi=session['total_ppi'])
    else:
        return render_template(
        'login.html',
        title='Login',
        year=datetime.now().year,
        message='Kindly Enter User Details'
        )

@db_bp.route('/filtered_data', methods=['POST'])
def filter_data():
    FILTER_DATA_DOCSTRING
    session['selected_ic'] = request.form.get('ic')
    session['selected_sbg'] = request.form.get('sbg')
    session['selected_bu'] = request.form.get('bu')
    session['selected_project'] = request.form.get('project')
    session['discipline_filter'] = request.form.get('discipline')
    session['data'] = chart(session['discipline_filter'],session['selected_project'],session['selected_bu'],session['selected_sbg'],session['selected_ic'])
    print(session['selected_project'])
    print(session['selected_ic'])
    

    return render_template('dashboard.html', processed_data=session['data'],ic_options=fetch_dropdown_options('ic'), bu=fetch_dropdown_options('bu'),
                           sbg=fetch_dropdown_options('sbg'), project=fetch_dropdown_options('project_name'),discipline=fetch_discipline_options('discipline'),
                           selected_ic=session['selected_ic'],selected_sbg=session['selected_sbg'], selected_bu=session['selected_bu'],
                           selected_project=session['selected_project'], selected_discipline=session['discipline_filter'],process_ppi=session['process_ppi'],mech_ppi=session['Mechanical_ppi'],elec_ppi=session['Electrical_ppi'],instru_ppi=session['Instrumentation_ppi'],total_ppi=session['total_ppi'])
    

@db_bp.route('/chart')
def chart_data():
    CHART_DATA_DOCSTRING
    return jsonify(session['data'])



@db_bp.route('/api/get_popup_data', methods=['POST'])
def popup_data():
    session['pop_data_dic'] = request.get_json()
    session['pop_discipline']= session['pop_data_dic'].get('discipline')
    session['pop_stage']= session['pop_data_dic'].get('stage')
    
    # print(session['pop_data_dic'])
    if session['pop_stage']==0:
        session['status_params_stage']=('Enquiry yet to float','')
    if session['pop_stage']==1:
        session['status_params_stage']=('Awaiting pre-offer from vendor', 'Pre-order pending with EDRC', 'Pre-order comments issued',
                        'Pre-order technically cleared', 'Pre-order technically rejected')
    if session['pop_stage']==2:
        session['status_params_stage']=('Sent to BU Operations',"") 
    if session['pop_stage']==3:
        session['status_params_stage']=('Post-order comments issued', 'Awaiting post-offer from vendor', 'Post-order pending with EDRC', 'Post-order technically cleared')
    if session['pop_stage']==4:
        session['status_params_stage']=('Approved by Client',"") 
    if session['pop_discipline']==0:
        session['status_params_disc']="Process"
    if session['pop_discipline']==1:
        session['status_params_disc']="Mechanical"
    if session['pop_discipline']==2:
        session['status_params_disc']="Electrical"
    if session['pop_discipline']==3:
        session['status_params_disc']="Instrumentation"
    if session['pop_discipline']==4:
        session['status_params_disc']="Others"
    
        
    conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')
    cursor = conn.cursor()
    session['popup_query']=(f'''SELECT project_name,discipline,equipment,item_vendor_code,vendor_name,status,elapsed_days FROM pr_data WHERE status IN {(session['status_params_stage'])} AND discipline= '{(session['status_params_disc'])}' ''')
    session['status_pop_up']=f'''SELECT status,item_code FROM pr_data  WHERE status IN {(session['status_params_stage'])} AND discipline= '{(session['status_params_disc'])}' '''
    session['average_time_popup']=f'''SELECT pr_issue_date_final,offer_recv_date,
                            comments_sent_pre,offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                            offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,
                            final_vendor_doc_pre,sent_to_scm,po_issued_date,
                            post_doc_rcv_dt,comments_sent_post,
                            final_vendor_doc_post,mfc_issued_date FROM pr_data  WHERE status IN {(session['status_params_stage'])} AND discipline= '{(session['status_params_disc'])}' '''
                            
    project_name_popup=session['selected_project']
# if session['selected_ic'] :  
    # print(session['selected_ic'])
    ic_popup=session['selected_ic'] 
# if session['selected_sbg']:
    sbg_popup=session['selected_sbg']
# if session['selected_bu']:
    bu_popup=session['selected_bu']
# if session['discipline_filter']:
    discipline_popup=session['discipline_filter']
    session["conditions_popup"] = []
    session["parameters_popup"] = []    
    if project_name_popup != 'Project' and project_name_popup != 'ALL' and project_name_popup is not None:
        print(project_name_popup )
        session['project_popup '] = project_name_popup 
        
        cursor.execute('SELECT job_code FROM pr_data WHERE project_name=?',(session['project_popup '],))
        
        job_code_popup  =cursor.fetchone()[0]
            
        session["conditions_popup"].append("job_code  = ?")
        session["parameters_popup"].append(job_code_popup )

    if ic_popup  != 'IC' and ic_popup  != 'ALL'   and ic_popup  is not None:
        session["conditions_popup"].append("ic = ?")
        session["parameters_popup"].append(ic_popup )

    if sbg_popup  != 'sbg' and  sbg_popup  != 'ALL'  and sbg_popup  is not None:
        session["conditions_popup"].append("sbg = ?")
        session["parameters_popup"].append(sbg_popup )

    if bu_popup  != 'BU' and  bu_popup  != 'ALL'  and bu_popup  is not None:
        session["conditions_popup"].append("bu = ?")
        session["parameters_popup"].append(bu_popup )
    if discipline_popup != 'Discipline' and  discipline_popup  != 'ALL'  and discipline_popup  is not None:
        session["conditions_popup"].append("discipline = ?")
        session["parameters_popup"].append(discipline_popup )

    if session["conditions_popup"]:
        session['popup_query'] += " AND " + " AND ".join(session["conditions_popup"]) 
        session['status_pop_up'] +=" AND " + " AND ".join(session["conditions_popup"])   
        session['average_time_popup'] +=" AND " + " AND ".join(session["conditions_popup"])   
        
        
        

    # print(session['popup_query']) 
       
    cursor.execute(session['popup_query'], tuple(session["parameters_popup"]))
    session["popup_query_data"] = cursor.fetchall()  
    cursor.execute(session['status_pop_up'], tuple(session["parameters_popup"]))
    session["status_elapsed_popup"] = cursor.fetchall()
    # print(session["status_elapsed_popup"])
    cursor.execute(session['average_time_popup'], tuple(session["parameters_popup"]))
    session["average_time_days_popup"] = cursor.fetchall()
    session['days_lapsed_popup']=elapsed_time(session["average_time_days_popup"],session['status_elapsed_popup'])
    
    for i, row in enumerate(session["popup_query_data"] ):
        row = list(row)
        row.pop(6)
        row.insert(6, session['days_lapsed_popup'][i])
        session["popup_query_data"][i] = row 
    return jsonify(session['popup_query_data'])

@db_bp.route('/download_help')
def download_help():
    DOWNLOAD_HELP_DOCSTRING
    return send_file('C:/Users/20323801/Desktop/VendorDesk/WWWSBG/Mainapp/User Manual/VE-User Manual.pdf', as_attachment=True)




@db_bp.route('/upload_excel', methods=['POST'])
def upload_excel():
    UPLOAD_EXCEL_DOCSTRING
    xl_file = request.files['file']
    if not xl_file:
        return jsonify({'error': f'No excel file selected. Select Excel file first!'})
    excel_file_path = 'C:/veuploads/updatepr/uploadedexcel.xlsx'
    xl_file.save(excel_file_path)
    conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')
    cursor = conn.cursor()
    if xl_file:
        session['df_updatepr'] = pd.read_excel(xl_file, engine='openpyxl', keep_default_na=False)
        missing_columns = ['DISCIPLINE', 'PROJECT', 'EQUIPMENT', 'PR NUMBER', 'PR VENDOR NUMBER']
        for col in missing_columns:
            if col not in session['df_updatepr'].columns:
                return jsonify({ 'error': f'Please check the template. Missing column: {col}'})
        for col in missing_columns:
            if len(session['df_updatepr'][[col]])==0:
                return jsonify({ 'error': f'Blank value found in column: {col}'})

        valid_discipline_options = ['Process', 'Mechanical', 'Electrical', 'Instrumentation', 'other']
        invalid_options = set(session['df_updatepr']['DISCIPLINE'].unique()) - set(valid_discipline_options)
        if invalid_options:
            return jsonify({ 'error': f'Invalid input found in DISCIPLINE column: {", ".join(invalid_options)}'})

        cursor.execute('SELECT project_name FROM project WHERE job_code=?', (session['selected_job_code'],))
        session['jobname'] = cursor.fetchone()
        if session['jobname']:
            for name in session['df_updatepr']['PROJECT']:
                if name!=session['jobname'][0]:
                      return jsonify({ 'error': f'Invalid input found in PROJECT column: {name}'})
        sql_dataframe = pd.read_sql_query('''SELECT project_name AS PROJECT,
       discipline AS DISCIPLINE,
       equipment AS EQUIPMENT,
       item_code AS 'PR NUMBER',
       item_vendor_code AS 'PR VENDOR NUMBER',
       vendor_name AS VENDOR,
       status AS STATUS,
       elapsed_days AS 'ELAPSED DAYS',
       remarks AS REMARKS,
       vendor_ph AS 'VENDOR CONTACT',
       vendor_mail AS 'VENDOR EMAIL',
       pr_issue_date_pre AS 'PR ISSUED DATE (PRELIMINARY BOQ)',
       pr_issue_date_final AS 'ENQUIRY SENT DATE',
       offer_recv_date AS 'OFFER RECEIVED DATE REV-A',
       comments_sent_pre AS 'COMMENTS SENT FROM EDRC DATE REV-A',
       offer_recv_date_2 AS 'OFFER RECEIVED DATE REV-B',
       comments_sent_pre_2 AS 'COMMENTS SENT FROM EDRC DATE REV-B',
       offer_recv_date_3 AS 'OFFER RECEIVED DATE REV-C',
       comments_sent_pre_3 AS 'COMMENTS SENT FROM EDRC DATE REV-C',
       offer_recv_date_4 AS 'OFFER RECEIVED DATE REV-D',
       comments_sent_pre_4 AS 'COMMENTS SENT FROM EDRC DATE REV-D',
       offer_recv_date_5 AS 'OFFER RECEIVED DATE REV-E',
       comments_sent_pre_5 AS 'COMMENTS SENT FROM EDRC DATE REV-E',
       tc_meeting_vendor_pre AS 'FINAL TECHNICAL CLEARANCE MEETING WITH VENDOR (PRE ORDER)',
       final_vendor_doc_pre AS 'FINAL DOCUMENT RECEIVED FROM VENDOR (PRE ORDER)',
       sent_to_scm AS 'Sent to BU Operations',
       priority AS 'PRIORITY (T1/T2/T3)',
       po_issued_date AS 'PO ISSUED DATE',
       finalized_vendor AS 'FINALIZED VENDOR',
       post_doc_rcv_dt AS 'POST ORDER DOC/DRG RECEIVED DATE',
       comments_sent_post AS 'COMMENTS SENT FROM EDRC DATE',
       tc_meeting_vendor_post AS 'FINAL TECHNICAL CLEARANCE MEETING WITH VENDOR (POST ORDER)',
       final_vendor_doc_post AS 'FINAL DOCUMENT RECEIVED FROM VENDOR (POST ORDER)',
       mfc_issued_date AS 'CLIENT APPROVAL DATE',
       target_completion_date AS 'PLANNED PRE-ORDER COMPLETION DATE',
       target_completion_date_po AS 'PLANNED PURCHASE-ORDER COMPLETION DATE',
       target_completion_date_post AS 'PLANNED POST-ORDER COMPLETION DATE',
       last_modified_by AS 'LAST MODIFIED BY'
FROM pr_data
WHERE job_code = ? 
''', conn, params=(session['selected_job_code'],))
                                  
        for index, row in session['df_updatepr'].iterrows():
            pr_vendor_number = row['PR VENDOR NUMBER']
            
            cursor.execute('SELECT * FROM pr_data WHERE item_vendor_code = ? AND job_code=?', (pr_vendor_number,session['selected_job_code'],))
            existing_row = cursor.fetchone() 
            if existing_row:
                i_f = sql_dataframe.index[sql_dataframe['PR VENDOR NUMBER'] == pr_vendor_number].tolist()
                if i_f:
                    differences = (session['df_updatepr'].iloc[[index]].reset_index(drop=True) != sql_dataframe.loc[i_f].reset_index(drop=True)).any(axis=1).any()
                    
                    if differences:
                        for column in sql_dataframe.loc[i_f].columns:
                            
                            sql_dataframe.loc[i_f,column] = session['df_updatepr'].iloc[index][column]
                else:
                    return jsonify({'error': f'PR VENDOR NUMBER {pr_vendor_number} not found in SQL database'})
            else:
                return jsonify({'error': f'New row detected for PR VENDOR NUMBER {pr_vendor_number}. Please add in "New PR Data tab"'})

        session['xldata_list'] = sql_dataframe.to_dict(orient='records')
        return jsonify({'columns': sql_dataframe.columns.tolist(), 'data': session['xldata_list']})



@db_bp.route('/top_sheet')
def top_sheet():
    conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')
    cursor = conn.cursor()
    
    query_A = """SELECT project_name AS 'Project Name', discipline AS Discipline, equipment AS Equipment, item_code AS 'PR Number', status AS Status FROM pr_data"""
    df_email = pd.read_sql_query(query_A, conn)
    query_B='''SELECT pr_issue_date_final,offer_recv_date,
                                comments_sent_pre,offer_recv_date_2,comments_sent_pre_2,offer_recv_date_3,comments_sent_pre_3,
                                offer_recv_date_4,comments_sent_pre_4,offer_recv_date_5,comments_sent_pre_5,
                                final_vendor_doc_pre,sent_to_scm,po_issued_date,
                                post_doc_rcv_dt,comments_sent_post,
                                final_vendor_doc_post,mfc_issued_date FROM pr_data'''
                                
    cursor.execute(query_B)
    dateFor = cursor.fetchall()                            
    query_c='''SELECT status,item_code FROM pr_data'''
    cursor.execute(query_c)
    meta = cursor.fetchall()             
  
    conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')           
    query='''SELECT 
    project_name,
    discipline,
    COUNT(DISTINCT item_code) AS Total,
    COUNT(DISTINCT CASE WHEN status = 'Enquiry yet to float' THEN item_vendor_code END) AS yetTOfloat,
        COUNT(DISTINCT CASE WHEN status IN (
                'Awaiting pre-offer from vendor', 
                'Pre-order pending with EDRC', 
                'Pre-order comments issued',
                'Pre-order technically cleared', 
                'Pre-order technically rejected', 
                'Regret offer',
                'Sent to BU Operations',
                'Post-order comments issued', 
                'Awaiting post-offer from vendor', 
                'Post-order pending with EDRC', 
                'Post-order technically cleared',
                'Approved by Client'
            ) THEN item_vendor_code END) ||
        "(" ||
        COUNT(DISTINCT CASE WHEN status IN (
                'Awaiting pre-offer from vendor', 
                'Pre-order pending with EDRC', 
                'Pre-order comments issued',
                'Pre-order technically cleared', 
                'Pre-order technically rejected', 
                'Regret offer',
                'Sent to BU Operations',
                'Post-order comments issued', 
                'Awaiting post-offer from vendor', 
                'Post-order pending with EDRC', 
                'Post-order technically cleared',
                'Approved by Client'
            ) THEN item_code END) ||
        ")"
     AS preEnquiry,
    COUNT(DISTINCT CASE WHEN status = 'Pre-order pending with EDRC' THEN item_vendor_code END) AS PendingWithEDRCPRE,
    COUNT(DISTINCT CASE WHEN status IN ('Awaiting pre-offer from vendor', 'Pre-order comments issued') THEN item_vendor_code END) AS PendingWithVendorPRE,
        COUNT(DISTINCT CASE WHEN status IN (
                'Pre-order technically cleared', 
                'Pre-order technically rejected', 
                'Regret offer', 
                'Sent to BU Operations', 
                'Approved by Client', 
                'Awaiting post-offer from vendor', 
                'Post-order comments issued', 
                'Post-order pending with EDRC', 
                'Post-order technically cleared'
            ) THEN item_vendor_code END) ||
        "(" ||
        COUNT(DISTINCT CASE WHEN status IN (
                'Pre-order technically cleared', 
                'Pre-order technically rejected', 
                'Regret offer', 
                'Sent to BU Operations', 
                'Approved by Client', 
                'Awaiting post-offer from vendor', 
                'Post-order comments issued', 
                'Post-order pending with EDRC', 
                'Post-order technically cleared'
            ) THEN item_code END) ||
        ")"
     AS total_pr_cleared,
    COUNT(DISTINCT CASE WHEN status = 'Sent to BU Operations' THEN item_vendor_code END) AS withSCM,
    COUNT(DISTINCT CASE WHEN status IN (
                'Post-order comments issued', 
                'Awaiting post-offer from vendor', 
                'Post-order pending with EDRC', 
                'Post-order technically cleared',
                'Approved by Client'
            ) THEN item_vendor_code END) AS withSCM,
    COUNT(DISTINCT CASE WHEN status IN ('Awaiting post-offer from vendor', 'Post-order comments issued') THEN item_vendor_code END) AS PendingWithVendorPOST,
    COUNT(DISTINCT CASE WHEN status = 'Post-order pending with EDRC' THEN item_vendor_code END) AS PendingWithEDRCPOST,
    COUNT(DISTINCT CASE WHEN status IN ('Post-order technically cleared', 'Approved by Client') THEN item_vendor_code END) AS Completed
FROM 
    pr_data 
GROUP BY 
    project_name, discipline

UNION ALL

SELECT 
    project_name,
    'Total' AS discipline,
    COUNT(DISTINCT item_code) AS Total,
    COUNT(DISTINCT CASE WHEN status = 'Enquiry yet to float' THEN item_vendor_code END) AS yetTOfloat,
     
        COUNT(DISTINCT CASE WHEN status IN (
                'Awaiting pre-offer from vendor', 
                'Pre-order pending with EDRC', 
                'Pre-order comments issued',
                'Pre-order technically cleared', 
                'Pre-order technically rejected', 
                'Regret offer',
                'Sent to BU Operations',
                'Post-order comments issued', 
                'Awaiting post-offer from vendor', 
                'Post-order pending with EDRC', 
                'Post-order technically cleared',
                'Approved by Client'
            ) THEN item_vendor_code END) ||
        "(" ||
        COUNT(DISTINCT CASE WHEN status IN (
                'Awaiting pre-offer from vendor', 
                'Pre-order pending with EDRC', 
                'Pre-order comments issued',
                'Pre-order technically cleared', 
                'Pre-order technically rejected', 
                'Regret offer',
                'Sent to BU Operations',
                'Post-order comments issued', 
                'Awaiting post-offer from vendor', 
                'Post-order pending with EDRC', 
                'Post-order technically cleared',
                'Approved by Client'
            ) THEN item_code END) ||
        ")"
     AS preEnquiry,
    COUNT(DISTINCT CASE WHEN status = 'Pre-order pending with EDRC' THEN item_vendor_code END) AS PendingWithEDRCPRE,
    COUNT(DISTINCT CASE WHEN status IN ('Awaiting pre-offer from vendor', 'Pre-order comments issued') THEN item_vendor_code END) AS PendingWithVendorPRE,
    
        COUNT(DISTINCT CASE WHEN status IN (
                'Pre-order technically cleared', 
                'Pre-order technically rejected', 
                'Regret offer', 
                'Sent to BU Operations', 
                'Approved by Client', 
                'Awaiting post-offer from vendor', 
                'Post-order comments issued', 
                'Post-order pending with EDRC', 
                'Post-order technically cleared'
            ) THEN item_vendor_code END) ||
        "(" ||
        COUNT(DISTINCT CASE WHEN status IN (
                'Pre-order technically cleared', 
                'Pre-order technically rejected', 
                'Regret offer', 
                'Sent to BU Operations', 
                'Approved by Client', 
                'Awaiting post-offer from vendor', 
                'Post-order comments issued', 
                'Post-order pending with EDRC', 
                'Post-order technically cleared'
            ) THEN item_code END) ||
        ")"
     AS total_pr_cleared,
    COUNT(DISTINCT CASE WHEN status = 'Sent to BU Operations' THEN item_vendor_code END) AS withSCM,
    COUNT(DISTINCT CASE WHEN status IN (
                'Post-order comments issued', 
                'Awaiting post-offer from vendor', 
                'Post-order pending with EDRC', 
                'Post-order technically cleared',
                'Approved by Client'
            ) THEN item_vendor_code END) AS Poissued,
            COUNT(DISTINCT CASE WHEN status = 'Post-order pending with EDRC' THEN item_vendor_code END) AS PendingWithEDRCPOST,
    COUNT(DISTINCT CASE WHEN status IN ('Awaiting post-offer from vendor', 'Post-order comments issued') THEN item_vendor_code END) AS PendingWithVendorPOST,
    
    COUNT(DISTINCT CASE WHEN status IN ('Post-order technically cleared', 'Approved by Client') THEN item_vendor_code END) AS Completed
FROM 
    pr_data 
GROUP BY 
    project_name

ORDER BY 
    project_name, discipline;
            '''                                
    df_summary = pd.read_sql_query(query, conn)
    conn.close()    

    APP_ROOT = os.path.dirname(os.path.abspath(__file__))
    xl_file_path = os.path.join(APP_ROOT, 'Top_Sheet.xlsx')
    xl_file_path1 = os.path.join(APP_ROOT, 'Top_Sheet-1.xlsx')
    wb = openpyxl.load_workbook(xl_file_path)
    ws = wb.active
    
     
    folpath = "C:\\WetdeskDownloads\\export\\topsheet"
    if "user" in session:
        userr = session["user"]
        session["folpath_user"] = folpath + "\\" + str(userr)
    
    
    
    for r_idx, row in enumerate(df_summary.values.tolist(), start=5):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')  
    wb.save(xl_file_path1)
   
        

    if not os.path.exists(session["folpath_user"]):
            os.makedirs(session["folpath_user"])
            print("NEW USER PATH: ",session["folpath_user"])
    file_list = os.listdir(session["folpath_user"])
    print(xl_file_path)
    for file_name in file_list:
        file_path = os.path.join(session["folpath_user"], file_name)
        try:
            # Delete the xl_file
            os.remove(file_path)
            print(f"Deleted: {file_path}")
        except Exception as e:
            print(f"Error deleting {file_path}: {e}")

    source_path = xl_file_path1
    print(source_path)
    session["excel_destination_path"] = session["folpath_user"] + "\\" + "Top_Sheet.xlsx"
    print(session["excel_destination_path"])

    if os.path.isfile(source_path):
        shutil.copy(source_path, session["excel_destination_path"])
    else:
        print(f'Error: {source_path} does not exist')
    p=session['selected_project'] 
    t = os.path.getctime(session["excel_destination_path"])
    t_str = time.ctime(t)    
    t_obj = time.strptime(t_str)  
    form_t = time.strftime("Top_Sheet-%Y-%m-%d-%H:%M:%S", t_obj)    
    form_t = form_t.replace(":", "-")   
    form_t = form_t.replace("-", "_") 
    os.rename(session["excel_destination_path"], os.path.split(session["excel_destination_path"])[0] + '/' + form_t + os.path.splitext(session["excel_destination_path"])[1])
    session["excel_destination_path"] = session["excel_destination_path"].replace("Top_Sheet",form_t)

    print(session["excel_destination_path"])
    
    return send_file(session["excel_destination_path"], as_attachment=True)
    
  
  

def send_otp_email(subject, body, recipient):
    try:
        outlook = win32com.client.Dispatch("Outlook.Application",pythoncom.CoInitialize())
        mail = outlook.CreateItem(0)
        mail.Subject = subject
        mail.BodyFormat = 2
        mail.HTMLBody = body
        mail.To = recipient
        mail.SentOnBehalfOfName = 'vendorenggedrc@lntecc.com'
        mail.Send()
        print(f"Email sent to {recipient}")
    except Exception as e:
        print("Error sending email:", e)

def generate_otp():
    return ''.join(random.choices(string.digits, k=6))
@db_bp.route('/send_otp', methods=['POST'])
def send_otp():
    email = request.form.get('email')
    if not email:
        return jsonify({'status': 'fail', 'message': 'Email is required.'}), 400

    otp = generate_otp()
    expiration_time = datetime.now() + timedelta(minutes=10)
    session['otp'] = otp
    session['otp_expiration'] = expiration_time.strftime('%Y-%m-%d %H:%M:%S')
    session['email'] = email

    subject = "Vendor Engineering Portal - OTP Notification"
    body = f"""
        <html>
        <body>
            <p>Dear Sir/Madam,</p>
            <p>Your OTP for resetting the password of Vendor Engineering portal is: <span style="font-weight:bold; color:blue">{session['otp']}.</span>
             This Password is valid for 10 minutes.</span></p>
            
            <p> <span style="font-style:italic; color:red">If you have not initated this process, please report to your Vendor Engineering Co-ordinator.</span></p>
            <p style="font-weight:bold">Regards,<br/>Vendor Engineering Cell<br/> WWW SBG EDRC</p>
            <p>--------------------------------------------------------------------------------------------------<br/>
            <span style="font-style:italic"> Note: This is a system-generated email from Vendor Engineering Portal. Please do not reply. </span> </p>
        </body>
        </html>
    """
    recipient = email

    send_otp_email(subject, body, recipient)

    return jsonify({'status': 'success', 'message': 'OTP sent to email!'})

@db_bp.route('/verify_otp', methods=['POST'])
def verify_otp():
    otp = request.form.get('otp')
    stored_otp = session.get('otp')
    expiration_time_str = session.get('otp_expiration')

    if not stored_otp or not expiration_time_str:
        return jsonify({'status': 'fail', 'message': 'OTP not found or expired'}), 400

    expiration_time = datetime.strptime(expiration_time_str, '%Y-%m-%d %H:%M:%S')
    if datetime.now() > expiration_time:
        return jsonify({'status': 'fail', 'message': 'OTP has expired!'}), 400

    if otp == stored_otp:
        session.pop('otp')
        session.pop('otp_expiration')
        session['otp_verified'] = True
        return jsonify({'status': 'success', 'message': 'OTP verified!'})
    else:
        return jsonify({'status': 'fail', 'message': 'Invalid OTP.'}), 400
    
@db_bp.route('/reset_password', methods=['POST'])
def reset_password():
   
    if not session.get('otp_verified'):
        print("no")
        return jsonify({'status': 'fail', 'message': 'OTP not verified!'}), 400

    email = session['email']
    print(email)
    new_password = request.form.get('new_password')
    hashed_password = generate_password_hash(new_password, method='sha256')

    conn = sqlite3.connect('C:\\Users\\Sayan\\Downloads\\Vendor_Engineering\\WETDESK2-Vendor_Engineering\\instance\\database.sqlite3')

    cursor = conn.cursor()
    cursor.execute('UPDATE user SET password = ? WHERE email = ?', (hashed_password, email))
    conn.commit()
    conn.close()

    session.pop('email')
    session.pop('otp_verified')

    return jsonify({'status': 'success', 'message': 'Password updated successfully!'})

